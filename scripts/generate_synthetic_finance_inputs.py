"""Generate deterministic synthetic financial inputs for the Finance AI Agent.

The generated workbooks model a small-to-medium university with four
departments. June contains deliberate, documented exceptions so anomaly
detection can be tested against realistic and internally reconciled data.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Iterable
from xml.sax.saxutils import escape

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable
from openpyxl.worksheet.table import TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table as PdfTable,
    TableStyle,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = PROJECT_ROOT / "data" / "synthetic"
MONTHLY_XLSX = OUTPUT_DIR / "monthly_financial_report_june_2026.xlsx"
ANNUAL_XLSX = OUTPUT_DIR / "annual_financial_report_2026.xlsx"
GOALS_PDF = OUTPUT_DIR / "financial_goals_2026.pdf"

DEPARTMENTS = ["Engineering", "Business", "Health Sciences", "Administration"]
MONTHS = pd.date_range("2026-01-01", "2026-12-01", freq="MS")
MONTH_NAMES = [month.strftime("%B") for month in MONTHS]

NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF2F8"
TEAL = "1F7A8C"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
AMBER = "FFC000"
LIGHT_AMBER = "FFF2CC"
RED = "C00000"
LIGHT_RED = "F4CCCC"
GRAY = "667085"
LIGHT_GRAY = "E7E6E6"
WHITE = "FFFFFF"
BLACK = "000000"
CURRENCY_FORMAT = '$#,##0;[Red]($#,##0);-'
PERCENT_FORMAT = '0.0%;[Red](0.0%);-'
INTEGER_FORMAT = '#,##0;[Red](#,##0);-'


@dataclass(frozen=True)
class FinanceData:
    """Purpose: hold all normalized report tables. Inputs: generated DataFrames. Outputs: immutable dataset."""

    revenue: pd.DataFrame
    expenses: pd.DataFrame
    payroll: pd.DataFrame
    students: pd.DataFrame
    cash_flow: pd.DataFrame
    scholarships: pd.DataFrame
    vendors: pd.DataFrame
    anomalies: pd.DataFrame


def allocate_integer(total: int, weights: Iterable[float]) -> list[int]:
    """Purpose: split an integer total exactly. Inputs: total and relative weights. Outputs: integer allocations summing to total."""

    normalized = [weight / sum(weights) for weight in weights]
    raw = [total * weight for weight in normalized]
    allocations = [int(value) for value in raw]
    # Assign rounding residue to the largest fractional remainders for exact tie-outs.
    residue = total - sum(allocations)
    order = sorted(range(len(raw)), key=lambda index: raw[index] - allocations[index], reverse=True)
    for index in order[:residue]:
        allocations[index] += 1
    return allocations


def build_revenue() -> pd.DataFrame:
    """Purpose: create monthly revenue detail. Inputs: fixed university assumptions. Outputs: revenue DataFrame."""

    annual_budget_by_department = {
        "Engineering": 9_200_000,
        "Business": 6_700_000,
        "Health Sciences": 8_100_000,
        "Administration": 2_000_000,
    }
    category_mix = {
        "Engineering": {"Tuition": 0.66, "Research Grants": 0.25, "Other Revenue": 0.09},
        "Business": {"Tuition": 0.83, "Research Grants": 0.07, "Other Revenue": 0.10},
        "Health Sciences": {"Tuition": 0.70, "Research Grants": 0.22, "Other Revenue": 0.08},
        "Administration": {"Service Fees": 0.55, "Donations": 0.25, "Other Revenue": 0.20},
    }
    seasonality = [0.075, 0.080, 0.090, 0.085, 0.083, 0.078, 0.070, 0.071, 0.100, 0.095, 0.088, 0.085]
    actual_factors = [0.995, 0.990, 0.982, 0.976, 0.968, 0.930, 0.957, 0.961, 0.969, 0.974, 0.978, 0.982]
    rows: list[dict[str, Any]] = []

    for department in DEPARTMENTS:
        department_monthly = allocate_integer(annual_budget_by_department[department], seasonality)
        for month_index, month in enumerate(MONTHS):
            categories = list(category_mix[department])
            category_budgets = allocate_integer(
                department_monthly[month_index],
                list(category_mix[department].values()),
            )
            for category_index, category in enumerate(categories):
                budget = category_budgets[category_index]
                factor = actual_factors[month_index]
                # Tuition underperformance is the main revenue exception; other sources remain near plan.
                if category == "Tuition":
                    factor -= 0.012
                elif category in {"Research Grants", "Donations"}:
                    factor += 0.020
                actual = round(budget * factor)
                rows.append(
                    {
                        "Period": month.date(),
                        "Month": month.strftime("%B"),
                        "Department": department,
                        "Revenue_Category": category,
                        "Budget_Revenue": budget,
                        "Actual_Revenue": actual,
                        "Variance": actual - budget,
                        "Variance_Pct": (actual - budget) / budget,
                    }
                )
    return pd.DataFrame(rows)


def build_expenses() -> pd.DataFrame:
    """Purpose: create monthly operating expense detail. Inputs: fixed budgets and anomaly rules. Outputs: expense DataFrame."""

    monthly_budget = {
        "Engineering": 610_000,
        "Business": 410_000,
        "Health Sciences": 555_000,
        "Administration": 370_000,
    }
    category_mix = {
        "Engineering": {"Payroll": 0.44, "Facilities": 0.16, "Equipment": 0.16, "Technology": 0.10, "Supplies": 0.08, "Services": 0.06},
        "Business": {"Payroll": 0.46, "Facilities": 0.18, "Equipment": 0.08, "Technology": 0.10, "Supplies": 0.07, "Services": 0.11},
        "Health Sciences": {"Payroll": 0.45, "Facilities": 0.18, "Equipment": 0.14, "Technology": 0.06, "Supplies": 0.12, "Services": 0.05},
        "Administration": {"Payroll": 0.47, "Facilities": 0.24, "Equipment": 0.04, "Technology": 0.08, "Supplies": 0.05, "Services": 0.12},
    }
    month_budget_factors = [0.96, 0.97, 0.99, 1.00, 1.01, 1.02, 0.97, 0.98, 1.03, 1.04, 1.02, 1.01]
    normal_actual_factors = [0.985, 0.990, 0.996, 1.005, 1.010, 1.018, 1.002, 1.006, 1.014, 1.019, 1.012, 1.008]
    rows: list[dict[str, Any]] = []

    for department in DEPARTMENTS:
        for month_index, month in enumerate(MONTHS):
            total_budget = round(monthly_budget[department] * month_budget_factors[month_index])
            categories = list(category_mix[department])
            category_budgets = allocate_integer(total_budget, list(category_mix[department].values()))
            actuals = [round(value * normal_actual_factors[month_index]) for value in category_budgets]

            if month.month == 6 and department == "Engineering":
                # Targeted category increases push Engineering to >15% over its June budget.
                for category, factor in {"Equipment": 1.48, "Facilities": 1.32, "Technology": 1.18, "Supplies": 1.16}.items():
                    actuals[categories.index(category)] = round(category_budgets[categories.index(category)] * factor)
                required = round(total_budget * 1.18)
                actuals[categories.index("Services")] += required - sum(actuals)

            if month.month == 6 and department == "Health Sciences":
                # The payroll line is synchronized with the payroll schedule created below.
                actuals[categories.index("Payroll")] = 342_000

            if month.month == 6 and department == "Administration":
                # The unusually high facilities invoice is visible both here and in vendor payments.
                actuals[categories.index("Facilities")] = 96_000

            for category_index, category in enumerate(categories):
                budget = category_budgets[category_index]
                actual = actuals[category_index]
                rows.append(
                    {
                        "Period": month.date(),
                        "Month": month.strftime("%B"),
                        "Department": department,
                        "Expense_Category": category,
                        "Budget_Expense": budget,
                        "Actual_Expense": actual,
                        "Variance": actual - budget,
                        "Variance_Pct": (actual - budget) / budget,
                    }
                )
    return pd.DataFrame(rows)


def build_payroll(expenses: pd.DataFrame) -> pd.DataFrame:
    """Purpose: derive payroll records that reconcile to expense payroll. Inputs: expense table. Outputs: payroll DataFrame."""

    headcount = {"Engineering": 76, "Business": 51, "Health Sciences": 68, "Administration": 48}
    rows: list[dict[str, Any]] = []
    payroll_expenses = expenses[expenses["Expense_Category"] == "Payroll"]

    for record in payroll_expenses.to_dict("records"):
        gross = int(record["Actual_Expense"])
        benefits = round(gross * 0.18)
        overtime = round(gross * (0.055 if record["Month"] == "June" and record["Department"] == "Health Sciences" else 0.018))
        base_salary = gross - benefits - overtime
        budget = int(record["Budget_Expense"])
        rows.append(
            {
                "Period": record["Period"],
                "Month": record["Month"],
                "Department": record["Department"],
                "Headcount_FTE": headcount[record["Department"]],
                "Base_Salary": base_salary,
                "Benefits": benefits,
                "Overtime": overtime,
                "Total_Payroll": gross,
                "Payroll_Budget": budget,
                "Variance": gross - budget,
                "Variance_Pct": (gross - budget) / budget,
            }
        )
    return pd.DataFrame(rows)


def build_student_payments(revenue: pd.DataFrame) -> pd.DataFrame:
    """Purpose: create invoice-level tuition collections. Inputs: tuition revenue totals. Outputs: student payment DataFrame."""

    academic_departments = DEPARTMENTS[:3]
    rows: list[dict[str, Any]] = []
    student_counter = 1001
    tuition = revenue[revenue["Revenue_Category"] == "Tuition"]

    for month in MONTHS:
        for department in academic_departments:
            tuition_row = tuition[
                (tuition["Period"] == month.date()) & (tuition["Department"] == department)
            ].iloc[0]
            billed_total = int(tuition_row["Budget_Revenue"])
            invoice_count = 16
            amounts = allocate_integer(billed_total, [1.0 + ((index % 5) * 0.08) for index in range(invoice_count)])
            overdue_indices = {1, 6, 11} if month.month == 6 else ({3} if month.month in {3, 9} else set())

            for index, amount_due in enumerate(amounts):
                due_date = date(2026, month.month, 12) + timedelta(days=index % 8)
                is_overdue = index in overdue_indices
                if is_overdue:
                    amount_paid = round(amount_due * (0.35 if index == 1 else 0.0))
                    payment_date = due_date + timedelta(days=37) if amount_paid else None
                    status = "Overdue - Partial" if amount_paid else "Overdue"
                else:
                    amount_paid = amount_due
                    payment_date = due_date - timedelta(days=(index % 4) + 1)
                    status = "Paid"
                rows.append(
                    {
                        "Invoice_ID": f"INV-{month.month:02d}-{department[:3].upper()}-{index + 1:03d}",
                        "Student_ID": f"STU-{student_counter:05d}",
                        "Department": department,
                        "Billing_Period": month.date(),
                        "Due_Date": due_date,
                        "Payment_Date": payment_date,
                        "Amount_Due": amount_due,
                        "Amount_Paid": amount_paid,
                        "Outstanding": amount_due - amount_paid,
                        "Status": status,
                        "Days_Overdue": 0 if status == "Paid" else 30 + index,
                    }
                )
                student_counter += 1
    return pd.DataFrame(rows)


def build_scholarships() -> pd.DataFrame:
    """Purpose: create scholarship allocation and disbursement detail. Inputs: policy assumptions. Outputs: scholarship DataFrame."""

    annual_allocations = {"Engineering": 880_000, "Business": 610_000, "Health Sciences": 790_000}
    type_mix = {"Merit": 0.40, "Need-Based": 0.45, "Research": 0.15}
    disbursement_factors = [0.92, 0.94, 0.96, 0.97, 0.98, 0.99, 0.95, 0.96, 1.00, 1.00, 0.98, 0.97]
    rows: list[dict[str, Any]] = []

    for department, annual_total in annual_allocations.items():
        monthly_allocations = allocate_integer(annual_total, [1] * 12)
        for month_index, month in enumerate(MONTHS):
            categories = list(type_mix)
            category_allocations = allocate_integer(monthly_allocations[month_index], list(type_mix.values()))
            for index, scholarship_type in enumerate(categories):
                allocated = category_allocations[index]
                awarded = round(allocated * disbursement_factors[month_index])
                rows.append(
                    {
                        "Period": month.date(),
                        "Month": month.strftime("%B"),
                        "Department": department,
                        "Scholarship_Type": scholarship_type,
                        "Allocated": allocated,
                        "Awarded": awarded,
                        "Remaining": allocated - awarded,
                        "Recipients": max(1, round(awarded / 4_800)),
                        "Utilization_Pct": awarded / allocated,
                    }
                )
    return pd.DataFrame(rows)


def build_vendor_payments(expenses: pd.DataFrame) -> pd.DataFrame:
    """Purpose: create vendor payment transactions tied to non-payroll expenses. Inputs: expense table. Outputs: vendor DataFrame."""

    vendor_by_category = {
        "Facilities": ("Northstar Facilities Group", "FAC"),
        "Equipment": ("Academic Equipment Partners", "EQP"),
        "Technology": ("Campus Technology Systems", "TEC"),
        "Supplies": ("University Supply Cooperative", "SUP"),
        "Services": ("Professional Services Network", "SRV"),
    }
    rows: list[dict[str, Any]] = []
    non_payroll = expenses[expenses["Expense_Category"] != "Payroll"]

    for record in non_payroll.to_dict("records"):
        total = int(record["Actual_Expense"])
        vendor_name, code = vendor_by_category[record["Expense_Category"]]
        month_number = pd.Timestamp(record["Period"]).month
        department_code = record["Department"][:3].upper()

        if month_number == 6 and record["Department"] == "Administration" and record["Expense_Category"] == "Facilities":
            parts = [96_000]
        elif month_number == 6 and record["Department"] == "Engineering" and record["Expense_Category"] == "Equipment":
            # Two identical payments intentionally share an invoice number and amount.
            duplicate_amount = 54_750
            parts = [duplicate_amount, duplicate_amount, total - (2 * duplicate_amount)]
        else:
            parts = allocate_integer(total, [0.58, 0.42])

        for index, amount in enumerate(parts):
            if amount <= 0:
                continue
            invoice_number = f"{code}-{month_number:02d}-{department_code}-{index + 1:02d}"
            if month_number == 6 and record["Department"] == "Engineering" and record["Expense_Category"] == "Equipment" and index in {0, 1}:
                invoice_number = "EQP-0626-ENG-4471"
            rows.append(
                {
                    "Payment_ID": f"PAY-{len(rows) + 1:05d}",
                    "Payment_Date": date(2026, month_number, min(25, 8 + index * 7)),
                    "Month": record["Month"],
                    "Department": record["Department"],
                    "Vendor": vendor_name,
                    "Invoice_Number": invoice_number,
                    "Expense_Category": record["Expense_Category"],
                    "Amount": amount,
                    "Payment_Method": "ACH",
                    "Approval_Status": "Approved",
                    "Potential_Duplicate": "Yes" if invoice_number == "EQP-0626-ENG-4471" else "No",
                    "High_Value_Flag": "Yes" if amount > 50_000 else "No",
                }
            )
    return pd.DataFrame(rows)


def build_cash_flow(revenue: pd.DataFrame, expenses: pd.DataFrame, scholarships: pd.DataFrame) -> pd.DataFrame:
    """Purpose: build a cash roll-forward. Inputs: revenue, expense, and scholarship tables. Outputs: cash flow DataFrame."""

    opening_cash = 5_850_000
    rows: list[dict[str, Any]] = []
    previous_actual = opening_cash
    previous_budget = opening_cash

    for month in MONTHS:
        month_date = month.date()
        revenue_budget = int(revenue.loc[revenue["Period"] == month_date, "Budget_Revenue"].sum())
        revenue_actual = int(revenue.loc[revenue["Period"] == month_date, "Actual_Revenue"].sum())
        expense_budget = int(expenses.loc[expenses["Period"] == month_date, "Budget_Expense"].sum())
        expense_actual = int(expenses.loc[expenses["Period"] == month_date, "Actual_Expense"].sum())
        scholarship_budget = int(scholarships.loc[scholarships["Period"] == month_date, "Allocated"].sum())
        scholarship_actual = int(scholarships.loc[scholarships["Period"] == month_date, "Awarded"].sum())
        capex_budget = 125_000 if month.month in {3, 6, 9, 12} else 80_000
        capex_actual = round(capex_budget * (1.35 if month.month == 6 else 1.02))
        budget_net = revenue_budget - expense_budget - scholarship_budget - capex_budget
        actual_net = revenue_actual - expense_actual - scholarship_actual - capex_actual
        ending_budget = previous_budget + budget_net
        ending_actual = previous_actual + actual_net
        rows.append(
            {
                "Period": month_date,
                "Month": month.strftime("%B"),
                "Beginning_Cash": previous_actual,
                "Budget_Cash_Inflows": revenue_budget,
                "Actual_Cash_Inflows": revenue_actual,
                "Budget_Operating_Outflows": expense_budget,
                "Actual_Operating_Outflows": expense_actual,
                "Budget_Scholarships": scholarship_budget,
                "Actual_Scholarships": scholarship_actual,
                "Budget_Capital_Outflows": capex_budget,
                "Actual_Capital_Outflows": capex_actual,
                "Budget_Net_Cash_Flow": budget_net,
                "Actual_Net_Cash_Flow": actual_net,
                "Budget_Ending_Cash": ending_budget,
                "Actual_Ending_Cash": ending_actual,
                "Ending_Cash_Variance": ending_actual - ending_budget,
            }
        )
        previous_actual = ending_actual
        previous_budget = ending_budget
    return pd.DataFrame(rows)


def build_anomalies(
    revenue: pd.DataFrame,
    expenses: pd.DataFrame,
    payroll: pd.DataFrame,
    students: pd.DataFrame,
    cash_flow: pd.DataFrame,
    vendors: pd.DataFrame,
) -> pd.DataFrame:
    """Purpose: document intentional exceptions with evidence. Inputs: report tables. Outputs: anomaly register."""

    june = date(2026, 6, 1)
    engineering_june = expenses[(expenses["Period"] == june) & (expenses["Department"] == "Engineering")]
    engineering_variance = engineering_june["Actual_Expense"].sum() / engineering_june["Budget_Expense"].sum() - 1
    health_may = payroll[(payroll["Month"] == "May") & (payroll["Department"] == "Health Sciences")]["Total_Payroll"].iloc[0]
    health_june = payroll[(payroll["Month"] == "June") & (payroll["Department"] == "Health Sciences")]["Total_Payroll"].iloc[0]
    overdue_count = int(((students["Billing_Period"] == june) & students["Status"].str.startswith("Overdue")).sum())
    tuition = revenue[revenue["Revenue_Category"] == "Tuition"]
    tuition_variance = tuition["Actual_Revenue"].sum() / tuition["Budget_Revenue"].sum() - 1
    june_cash = cash_flow[cash_flow["Period"] == june].iloc[0]
    duplicate_amount = vendors.loc[vendors["Invoice_Number"] == "EQP-0626-ENG-4471", "Amount"].iloc[0]

    rows = [
        ("ANOM-001", june, "Engineering", "Department Overspending", "High", "Actual expenses exceed June budget by more than 15%.", engineering_variance, 0.12, "Budget_vs_Actual"),
        ("ANOM-002", june, "Health Sciences", "Payroll Spike", "High", "June payroll increased sharply relative to May.", health_june / health_may - 1, 0.10, "Payroll"),
        ("ANOM-003", june, "Multiple", "Overdue Student Payments", "Medium", f"{overdue_count} June tuition invoices are overdue or partially paid.", overdue_count, 8, "Student_Payments"),
        ("ANOM-004", june, "Engineering", "Duplicate Vendor Payment", "Critical", "Two payments share the same vendor, invoice number, and amount.", duplicate_amount, 0, "Vendor_Payments"),
        ("ANOM-005", june, "Administration", "Unusually High Facilities Expense", "High", "Single facilities payment materially exceeds normal transaction size.", 96_000, 50_000, "Vendor_Payments"),
        ("ANOM-006", date(2026, 12, 31), "University-wide", "Tuition Revenue Below Target", "High", "Annual tuition revenue remains below budget.", tuition_variance, -0.02, "Revenue"),
        ("ANOM-007", june, "University-wide", "Cash Flow Below Expectation", "High", "June net cash flow is materially below budget.", june_cash["Actual_Net_Cash_Flow"] - june_cash["Budget_Net_Cash_Flow"], 0, "Cash_Flow"),
    ]
    return pd.DataFrame(
        rows,
        columns=["Anomaly_ID", "Detected_Period", "Department", "Anomaly_Type", "Severity", "Description", "Observed_Value", "Threshold", "Source_Sheet"],
    )


def generate_finance_data() -> FinanceData:
    """Purpose: orchestrate all internally linked datasets. Inputs: none. Outputs: complete FinanceData bundle."""

    revenue = build_revenue()
    expenses = build_expenses()
    payroll = build_payroll(expenses)
    students = build_student_payments(revenue)
    scholarships = build_scholarships()
    vendors = build_vendor_payments(expenses)
    cash_flow = build_cash_flow(revenue, expenses, scholarships)
    anomalies = build_anomalies(revenue, expenses, payroll, students, cash_flow, vendors)
    return FinanceData(revenue, expenses, payroll, students, cash_flow, scholarships, vendors, anomalies)


def filter_data(data: FinanceData, monthly: bool) -> FinanceData:
    """Purpose: select June-only or full-year records. Inputs: FinanceData and monthly flag. Outputs: filtered FinanceData."""

    if not monthly:
        return data
    june = date(2026, 6, 1)
    return FinanceData(
        revenue=data.revenue[data.revenue["Period"] == june].copy(),
        expenses=data.expenses[data.expenses["Period"] == june].copy(),
        payroll=data.payroll[data.payroll["Period"] == june].copy(),
        students=data.students[data.students["Billing_Period"] == june].copy(),
        cash_flow=data.cash_flow[data.cash_flow["Period"] == june].copy(),
        scholarships=data.scholarships[data.scholarships["Period"] == june].copy(),
        vendors=data.vendors[pd.to_datetime(data.vendors["Payment_Date"]).dt.month == 6].copy(),
        anomalies=data.anomalies[
            (data.anomalies["Detected_Period"] == june)
            | (data.anomalies["Anomaly_Type"] == "Tuition Revenue Below Target")
        ].copy(),
    )


def build_department_summary(data: FinanceData) -> pd.DataFrame:
    """Purpose: aggregate financial performance by department. Inputs: filtered FinanceData. Outputs: department summary."""

    revenue = data.revenue.groupby("Department")[["Budget_Revenue", "Actual_Revenue"]].sum()
    expenses = data.expenses.groupby("Department")[["Budget_Expense", "Actual_Expense"]].sum()
    result = revenue.join(expenses, how="outer").fillna(0).reset_index()
    result["Net_Contribution"] = result["Actual_Revenue"] - result["Actual_Expense"]
    result["Expense_Variance"] = result["Actual_Expense"] - result["Budget_Expense"]
    result["Expense_Variance_Pct"] = result["Expense_Variance"] / result["Budget_Expense"]
    result["Status"] = result["Expense_Variance_Pct"].apply(
        lambda value: "Flag - Overspend" if value > 0.12 else ("Watch" if value > 0.08 else "Within Threshold")
    )
    return result


def build_budget_vs_actual(data: FinanceData) -> pd.DataFrame:
    """Purpose: combine revenue and expense variance views. Inputs: filtered FinanceData. Outputs: variance table."""

    group_columns = ["Period", "Month", "Department"]
    revenue = data.revenue.groupby(group_columns, as_index=False)[["Budget_Revenue", "Actual_Revenue"]].sum()
    expenses = data.expenses.groupby(group_columns, as_index=False)[["Budget_Expense", "Actual_Expense"]].sum()
    result = revenue.merge(expenses, on=group_columns, how="outer").fillna(0)
    result["Revenue_Variance"] = result["Actual_Revenue"] - result["Budget_Revenue"]
    result["Revenue_Variance_Pct"] = result["Revenue_Variance"] / result["Budget_Revenue"]
    result["Expense_Variance"] = result["Actual_Expense"] - result["Budget_Expense"]
    result["Expense_Variance_Pct"] = result["Expense_Variance"] / result["Budget_Expense"]
    result["Net_Budget"] = result["Budget_Revenue"] - result["Budget_Expense"]
    result["Net_Actual"] = result["Actual_Revenue"] - result["Actual_Expense"]
    return result


def build_executive_summary(data: FinanceData, monthly: bool) -> pd.DataFrame:
    """Purpose: calculate decision-ready headline KPIs. Inputs: filtered FinanceData and period flag. Outputs: KPI table."""

    total_revenue_budget = data.revenue["Budget_Revenue"].sum()
    total_revenue_actual = data.revenue["Actual_Revenue"].sum()
    total_expense_budget = data.expenses["Budget_Expense"].sum()
    total_expense_actual = data.expenses["Actual_Expense"].sum()
    total_payroll = data.payroll["Total_Payroll"].sum()
    billed = data.students["Amount_Due"].sum()
    paid = data.students["Amount_Paid"].sum()
    overdue = data.students["Status"].str.startswith("Overdue").sum()
    ending_cash = data.cash_flow.iloc[-1]["Actual_Ending_Cash"]
    monthly_operating_expense = total_expense_actual if monthly else total_expense_actual / 12
    reserve_months = ending_cash / monthly_operating_expense
    metrics = [
        ("Actual Revenue", total_revenue_actual, total_revenue_budget, total_revenue_actual - total_revenue_budget, "Currency", "Below target" if total_revenue_actual < total_revenue_budget else "On target"),
        ("Actual Operating Expenses", total_expense_actual, total_expense_budget, total_expense_actual - total_expense_budget, "Currency", "Over budget" if total_expense_actual > total_expense_budget else "Within budget"),
        ("Operating Surplus", total_revenue_actual - total_expense_actual, total_revenue_budget - total_expense_budget, (total_revenue_actual - total_expense_actual) - (total_revenue_budget - total_expense_budget), "Currency", "Monitor"),
        ("Payroll as % of Revenue", total_payroll / total_revenue_actual, 0.42, total_payroll / total_revenue_actual - 0.42, "Percent", "Within goal" if total_payroll / total_revenue_actual < 0.42 else "Flag"),
        ("Tuition Collection Rate", paid / billed, 0.94, paid / billed - 0.94, "Percent", "Within goal" if paid / billed > 0.94 else "Flag"),
        ("Overdue Payment Rate", overdue / len(data.students), 0.06, overdue / len(data.students) - 0.06, "Percent", "Flag" if overdue / len(data.students) > 0.06 else "Within goal"),
        ("Ending Cash", ending_cash, data.cash_flow.iloc[-1]["Budget_Ending_Cash"], data.cash_flow.iloc[-1]["Ending_Cash_Variance"], "Currency", "Below plan" if data.cash_flow.iloc[-1]["Ending_Cash_Variance"] < 0 else "On plan"),
        ("Cash Reserve (months)", reserve_months, 3.0, reserve_months - 3.0, "Number", "Within goal" if reserve_months >= 3 else "Flag"),
        ("Open Anomalies", len(data.anomalies), 0, len(data.anomalies), "Count", "Review required"),
    ]
    return pd.DataFrame(metrics, columns=["Metric", "Actual", "Goal_or_Budget", "Variance", "Format", "Status"])


def write_dataframe(writer: pd.ExcelWriter, sheet_name: str, frame: pd.DataFrame) -> None:
    """Purpose: write a report table through pandas. Inputs: writer, sheet name, and DataFrame. Outputs: populated worksheet."""

    frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)


def apply_table_style(worksheet: Any, frame: pd.DataFrame, table_name: str) -> None:
    """Purpose: apply professional table and column formatting. Inputs: worksheet, DataFrame, table name. Outputs: styled worksheet."""

    start_row = 5
    end_row = start_row + len(frame)
    end_column = len(frame.columns)
    if len(frame):
        table = ExcelTable(displayName=table_name, ref=f"A{start_row}:{get_column_letter(end_column)}{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    thin_gray = Side(style="thin", color="D0D5DD")
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A6"
    worksheet.auto_filter.ref = f"A{start_row}:{get_column_letter(end_column)}{end_row}"
    worksheet.row_dimensions[start_row].height = 30

    for cell in worksheet[start_row]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows(min_row=start_row + 1, max_row=end_row):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=BLACK)
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin_gray)

    for column_index, column_name in enumerate(frame.columns, 1):
        series = frame[column_name]
        column_letter = get_column_letter(column_index)
        header_width = len(str(column_name).replace("_", " ")) + 2
        if pd.api.types.is_datetime64_any_dtype(series) or "Date" in column_name or "Period" in column_name:
            width = max(13, header_width)
            for cell in worksheet[column_letter][start_row:]:
                cell.number_format = "mmm d, yyyy"
        elif any(token in column_name for token in ["Pct", "Rate", "Utilization"]):
            width = max(13, header_width)
            for cell in worksheet[column_letter][start_row:]:
                cell.number_format = PERCENT_FORMAT
        elif any(token in column_name for token in ["Revenue", "Expense", "Cash", "Salary", "Benefits", "Overtime", "Payroll", "Amount", "Allocated", "Awarded", "Remaining", "Variance", "Contribution", "Outstanding"]):
            width = max(15, header_width)
            for cell in worksheet[column_letter][start_row:]:
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")
        elif pd.api.types.is_numeric_dtype(series):
            width = max(12, header_width)
            for cell in worksheet[column_letter][start_row:]:
                cell.number_format = INTEGER_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            sample_width = max([len(str(value)) for value in series.head(60)] + [header_width])
            width = min(max(12, sample_width + 2), 44)
        worksheet.column_dimensions[column_letter].width = min(width, 44)


def add_sheet_heading(worksheet: Any, title: str, subtitle: str, column_count: int) -> None:
    """Purpose: add consistent report titles and metadata. Inputs: worksheet, text, and span. Outputs: formatted heading."""

    end_column = max(4, column_count)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    worksheet["A1"] = title
    worksheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    worksheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 34
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    worksheet["A2"] = subtitle
    worksheet["A2"].font = Font(name="Aptos", size=10, italic=True, color=GRAY)
    worksheet["A2"].alignment = Alignment(vertical="center")
    worksheet["A3"] = "Currency: USD | Synthetic data | Generated deterministically"
    worksheet["A3"].font = Font(name="Aptos", size=9, color=GRAY)


def format_executive_summary(worksheet: Any, frame: pd.DataFrame, period_label: str) -> None:
    """Purpose: style KPI dashboard and add charts. Inputs: worksheet, KPI table, and period label. Outputs: executive sheet."""

    add_sheet_heading(worksheet, f"University Financial Executive Summary - {period_label}", "Management reporting view with goals, variances, and exception status", len(frame.columns))
    worksheet.freeze_panes = "A6"
    format_map = {row + 6: value for row, value in enumerate(frame["Format"])}
    for row_index, format_name in format_map.items():
        for column in ["B", "C", "D"]:
            worksheet[f"{column}{row_index}"].number_format = (
                PERCENT_FORMAT if format_name == "Percent" else CURRENCY_FORMAT if format_name == "Currency" else "0.0"
            )

    status_column = frame.columns.get_loc("Status") + 1
    status_range = f"{get_column_letter(status_column)}6:{get_column_letter(status_column)}{5 + len(frame)}"
    worksheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Flag"'], fill=PatternFill("solid", fgColor=LIGHT_RED)),
    )
    worksheet.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Within goal"'], fill=PatternFill("solid", fgColor=LIGHT_GREEN)),
    )


def add_annual_charts(workbook: Any) -> None:
    """Purpose: add concise monthly trend charts. Inputs: annual workbook. Outputs: chart-enhanced summary sheets."""

    budget_sheet = workbook["Budget_vs_Actual"]
    max_row = budget_sheet.max_row
    chart = LineChart()
    chart.title = "Monthly Revenue: Budget vs Actual"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "Month / Department"
    chart.height = 7.5
    chart.width = 16
    data = Reference(budget_sheet, min_col=5, max_col=6, min_row=5, max_row=max_row)
    categories = Reference(budget_sheet, min_col=2, min_row=6, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.style = 13
    budget_sheet.add_chart(chart, "N5")

    cash_sheet = workbook["Cash_Flow"]
    cash_chart = LineChart()
    cash_chart.title = "Ending Cash: Budget vs Actual"
    cash_chart.y_axis.title = "USD"
    cash_chart.x_axis.title = "Month"
    cash_chart.height = 7.5
    cash_chart.width = 16
    cash_data = Reference(cash_sheet, min_col=14, max_col=15, min_row=5, max_row=cash_sheet.max_row)
    cash_categories = Reference(cash_sheet, min_col=2, min_row=6, max_row=cash_sheet.max_row)
    cash_chart.add_data(cash_data, titles_from_data=True)
    cash_chart.set_categories(cash_categories)
    cash_chart.style = 13
    cash_sheet.add_chart(cash_chart, "R5")


def apply_exception_formatting(workbook: Any) -> None:
    """Purpose: highlight thresholds and anomalies. Inputs: workbook. Outputs: conditional formatting rules."""

    for sheet_name in ["Revenue", "Expenses", "Budget_vs_Actual", "Department_Summary", "Payroll"]:
        worksheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in worksheet[5]}
        for header in ["Variance_Pct", "Expense_Variance_Pct"]:
            if header in headers:
                letter = get_column_letter(headers[header])
                target = f"{letter}6:{letter}{worksheet.max_row}"
                worksheet.conditional_formatting.add(
                    target,
                    ColorScaleRule(
                        start_type="num",
                        start_value=-0.12,
                        start_color=LIGHT_GREEN,
                        mid_type="num",
                        mid_value=0,
                        mid_color=WHITE,
                        end_type="num",
                        end_value=0.15,
                        end_color=LIGHT_RED,
                    ),
                )

    anomalies = workbook["Anomalies_Embedded"]
    severity_column = {cell.value: cell.column for cell in anomalies[5]}["Severity"]
    letter = get_column_letter(severity_column)
    for severity, fill in [("Critical", RED), ("High", "F4B183"), ("Medium", LIGHT_AMBER)]:
        anomalies.conditional_formatting.add(
            f"{letter}6:{letter}{anomalies.max_row}",
            CellIsRule(operator="equal", formula=[f'"{severity}"'], fill=PatternFill("solid", fgColor=fill)),
        )


def create_workbook(data: FinanceData, output_path: Path, monthly: bool) -> None:
    """Purpose: produce a polished Excel financial report. Inputs: dataset, path, and period mode. Outputs: saved XLSX."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_data = filter_data(data, monthly)
    period_label = "June 2026" if monthly else "Fiscal Year 2026"
    frames = {
        "Executive_Summary": build_executive_summary(report_data, monthly),
        "Revenue": report_data.revenue,
        "Expenses": report_data.expenses,
        "Budget_vs_Actual": build_budget_vs_actual(report_data),
        "Department_Summary": build_department_summary(report_data),
        "Payroll": report_data.payroll,
        "Student_Payments": report_data.students,
        "Cash_Flow": report_data.cash_flow,
        "Scholarships": report_data.scholarships,
        "Vendor_Payments": report_data.vendors,
        "Anomalies_Embedded": report_data.anomalies,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            write_dataframe(writer, sheet_name, frame)

    workbook = load_workbook(output_path)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    for index, (sheet_name, frame) in enumerate(frames.items(), 1):
        worksheet = workbook[sheet_name]
        add_sheet_heading(
            worksheet,
            f"{sheet_name.replace('_', ' ')} - {period_label}",
            "Small/medium university | Engineering, Business, Health Sciences, and Administration",
            len(frame.columns),
        )
        apply_table_style(worksheet, frame, f"Table{index:02d}{sheet_name.replace('_', '')[:15]}")
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.oddFooter.center.text = f"{period_label} | Confidential synthetic test data"
        worksheet.oddFooter.right.text = "Page &P of &N"
        worksheet.auto_filter.ref = None  # The Excel table already supplies the filter UI.

    format_executive_summary(workbook["Executive_Summary"], frames["Executive_Summary"], period_label)
    apply_exception_formatting(workbook)
    if not monthly:
        add_annual_charts(workbook)
    workbook.active = 0
    workbook.save(output_path)


def pdf_header_footer(canvas: Any, document: Any) -> None:
    """Purpose: draw recurring PDF header/footer. Inputs: ReportLab canvas and document. Outputs: decorated page."""

    canvas.saveState()
    width, height = letter
    canvas.setFillColor(colors.HexColor(f"#{NAVY}"))
    canvas.rect(0, height - 0.42 * inch, width, 0.42 * inch, fill=True, stroke=False)
    canvas.setFont("Helvetica-Bold", 8)
    canvas.setFillColor(colors.white)
    canvas.drawString(0.65 * inch, height - 0.27 * inch, "FINANCE AI AGENT | UNIVERSITY FINANCIAL GOVERNANCE")
    canvas.setFillColor(colors.HexColor("#667085"))
    canvas.setFont("Helvetica", 8)
    canvas.drawString(0.65 * inch, 0.38 * inch, "Synthetic policy input | Fiscal Year 2026")
    canvas.drawRightString(width - 0.65 * inch, 0.38 * inch, f"Page {document.page}")
    canvas.restoreState()


def build_goal_table(data: list[list[str]], widths: list[float]) -> PdfTable:
    """Purpose: create a consistent PDF policy table. Inputs: cell data and column widths. Outputs: styled ReportLab Table."""

    header_cell_style = ParagraphStyle(
        "GoalTableHeader",
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=10.5,
        textColor=colors.white,
    )
    body_cell_style = ParagraphStyle(
        "GoalTableBody",
        fontName="Helvetica",
        fontSize=8.5,
        leading=10.8,
        textColor=colors.HexColor("#101828"),
    )
    # Paragraph cells wrap naturally inside their assigned column widths.
    wrapped_data = [
        [
            Paragraph(escape(str(value)), header_cell_style if row_index == 0 else body_cell_style)
            for value in row
        ]
        for row_index, row in enumerate(data)
    ]
    table = PdfTable(wrapped_data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("LEADING", (0, 0), (-1, -1), 11),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D0D5DD")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def create_goals_pdf(output_path: Path) -> None:
    """Purpose: create the professional FY2026 financial goals document. Inputs: output path. Outputs: saved PDF."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleCustom",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=24,
        leading=29,
        textColor=colors.HexColor(f"#{NAVY}"),
        alignment=TA_LEFT,
        spaceAfter=12,
    )
    subtitle_style = ParagraphStyle(
        "SubtitleCustom",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#475467"),
        spaceAfter=20,
    )
    heading_style = ParagraphStyle(
        "HeadingCustom",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=colors.HexColor(f"#{NAVY}"),
        spaceBefore=9,
        spaceAfter=7,
    )
    body_style = ParagraphStyle(
        "BodyCustom",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=14,
        textColor=colors.HexColor("#344054"),
        alignment=TA_LEFT,
        spaceAfter=8,
    )
    callout_style = ParagraphStyle(
        "Callout",
        parent=body_style,
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor(f"#{NAVY}"),
        borderColor=colors.HexColor(f"#{BLUE}"),
        borderWidth=1,
        borderPadding=9,
        backColor=colors.HexColor("#EAF2F8"),
        spaceAfter=12,
    )

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        rightMargin=0.65 * inch,
        leftMargin=0.65 * inch,
        topMargin=0.68 * inch,
        bottomMargin=0.65 * inch,
        title="Financial Goals 2026",
        author="Finance AI Agent Project",
        subject="Synthetic university financial goals and monitoring rules",
    )
    story: list[Any] = [
        Spacer(1, 0.24 * inch),
        Paragraph("Financial Goals 2026", title_style),
        Paragraph("Small/Medium University | Governance thresholds for automated financial analysis", subtitle_style),
        Paragraph(
            "Purpose: provide measurable, machine-readable financial goals against which monthly and annual reports can be validated, compared, and explained. All financial calculations remain deterministic; AI-generated analysis should interpret the results without recalculating them.",
            callout_style,
        ),
        Paragraph("Executive Overview", heading_style),
        Paragraph(
            "The university will balance academic investment with disciplined cost control and liquidity protection. Leadership will prioritize tuition collection, payroll sustainability, department accountability, and early identification of unusual transactions. Exceptions require documented review, not automatic rejection.",
            body_style,
        ),
        Paragraph("Core Annual Goals", heading_style),
        build_goal_table(
            [
                ["Area", "FY2026 Goal", "Measurement and Action"],
                ["Annual revenue", "$26.0 million budget; achieve at least 98% of total revenue target", "Compare actual revenue with approved budget monthly and year-to-date."],
                ["Expense control", "Keep total operating expenses at or below approved budget", "Investigate material unfavorable category and department variances."],
                ["Payroll cost", "Below 42% of total revenue", "Flag when monthly or year-to-date payroll ratio is 42% or higher."],
                ["Tuition collection", "Collection rate above 94%", "Amount paid divided by amount due; review aging and payment plans."],
                ["Scholarships", "Deploy 95%-100% of approved allocation", "Protect access goals while avoiding unapproved over-allocation."],
                ["Cash reserve", "At least 3 months of operating expenses", "Ending unrestricted cash divided by average monthly operating expense."],
            ],
            [1.35 * inch, 2.25 * inch, 3.35 * inch],
        ),
        Spacer(1, 0.12 * inch),
        Paragraph("Department Budget Variance Thresholds", heading_style),
        Paragraph(
            "Department leaders are accountable for approved budgets and documented corrective action. Normal operating tolerance is within +/-8% of budget. A variance between 8% and 12% is a watch item. More than 12% overspending is a mandatory financial risk flag.",
            body_style,
        ),
        build_goal_table(
            [
                ["Variance", "Classification", "Required Response"],
                ["Within +/-8%", "Acceptable", "Monitor through standard monthly reporting."],
                ["8% to 12% overspend", "Watch", "Department explanation and near-term forecast update."],
                [">12% overspend", "Flag", "Immediate review by Finance; corrective action and approval assessment."],
            ],
            [1.45 * inch, 1.35 * inch, 4.15 * inch],
        ),
        PageBreak(),
        Paragraph("Revenue and Tuition Collection Goals", heading_style),
        Paragraph(
            "Revenue reporting must separate tuition, research grants, service fees, donations, and other revenue. Tuition performance is measured against both billed value and cash collected. The annual tuition collection rate must exceed 94%, while overdue payments must remain below 6% of open and issued student invoices.",
            body_style,
        ),
        Paragraph("Expense and Payroll Cost Goals", heading_style),
        Paragraph(
            "Operating expense analysis must compare budget and actual values by month, department, and expense category. Payroll includes base salary, benefits, and overtime. Payroll must remain below 42% of total revenue, and unusual overtime or month-over-month changes require explanation.",
            body_style,
        ),
        Paragraph("Scholarship Allocation Goals", heading_style),
        Paragraph(
            "Scholarship reporting must reconcile approved allocation, awarded value, remaining balance, and recipient count by department and scholarship type. The target utilization range is 95%-100% by year-end. Awards may not exceed the approved allocation without formal authorization.",
            body_style,
        ),
        Paragraph("Cash Reserve Goals", heading_style),
        Paragraph(
            "The university will maintain unrestricted cash equal to at least three months of average operating expenses. Monthly cash flow reporting must reconcile beginning cash, inflows, operating outflows, scholarship disbursements, capital outflows, net cash flow, and ending cash. Negative variance to the cash plan requires commentary.",
            body_style,
        ),
        Paragraph("Risk Monitoring Rules", heading_style),
        build_goal_table(
            [
                ["Rule", "Threshold", "Flagging Standard"],
                ["Department overspending", "More than 12% above budget", "High-priority department variance flag."],
                ["Vendor payment", "Above $50,000", "High-value transaction review; validate approval and support."],
                ["Duplicate payment", "Same vendor + invoice + amount", "Critical exception; investigate before next payment cycle."],
                ["Expense acceleration", "Month-over-month increase above 10%", "Identify category, department, and business driver."],
                ["Overdue student payments", "6% or more", "Collection risk flag and aging review."],
                ["Payroll ratio", "42% or more of total revenue", "Cost sustainability flag."],
                ["Cash reserve", "Below 3 months", "Liquidity flag with cash recovery plan."],
            ],
            [1.55 * inch, 1.75 * inch, 3.65 * inch],
        ),
        Spacer(1, 0.12 * inch),
        Paragraph("Monthly Reporting Standards", heading_style),
        build_goal_table(
            [
                ["Standard", "Requirement"],
                ["Reporting deadline", "Issue the management report within 10 business days after month-end."],
                ["Required comparisons", "Actual vs budget, variance dollars, variance percent, prior month, and year-to-date context."],
                ["Required dimensions", "Department, revenue/expense category, vendor, payroll component, and student payment status."],
                ["Reconciliation", "Revenue, expenses, payroll, scholarships, vendor payments, and cash flow must tie to their reported totals."],
                ["Exception commentary", "Every high or critical flag must identify cause, owner, response, and expected resolution date."],
                ["Data quality", "No missing reporting periods, duplicate identifiers, unexplained negative values, or unresolved total mismatches."],
            ],
            [1.75 * inch, 5.2 * inch],
        ),
        Spacer(1, 0.18 * inch),
        Paragraph(
            "Approval and use: These goals are synthetic project inputs designed to exercise the Finance AI Agent's deterministic calculation, anomaly detection, and report interpretation workflows.",
            callout_style,
        ),
    ]
    document.build(story, onFirstPage=pdf_header_footer, onLaterPages=pdf_header_footer)


def validate_outputs(data: FinanceData) -> None:
    """Purpose: assert core reconciliation and anomaly conditions. Inputs: complete FinanceData. Outputs: raises on failure."""

    june = date(2026, 6, 1)
    engineering = data.expenses[(data.expenses["Period"] == june) & (data.expenses["Department"] == "Engineering")]
    assert engineering["Actual_Expense"].sum() / engineering["Budget_Expense"].sum() - 1 > 0.15
    health = data.payroll[data.payroll["Department"] == "Health Sciences"].set_index("Month")
    assert health.loc["June", "Total_Payroll"] / health.loc["May", "Total_Payroll"] - 1 > 0.10
    assert (data.vendors["Invoice_Number"] == "EQP-0626-ENG-4471").sum() == 2
    assert data.vendors.loc[data.vendors["Invoice_Number"] == "EQP-0626-ENG-4471", "Amount"].nunique() == 1
    assert ((data.students["Billing_Period"] == june) & data.students["Status"].str.startswith("Overdue")).sum() >= 6
    assert data.revenue.loc[data.revenue["Revenue_Category"] == "Tuition", "Actual_Revenue"].sum() < data.revenue.loc[data.revenue["Revenue_Category"] == "Tuition", "Budget_Revenue"].sum()

    # Payroll and vendor schedules must reconcile exactly to their parent expense lines.
    payroll_totals = data.payroll.groupby(["Period", "Department"])["Total_Payroll"].sum().sort_index()
    expense_payroll = data.expenses[data.expenses["Expense_Category"] == "Payroll"].groupby(["Period", "Department"])["Actual_Expense"].sum().sort_index()
    pd.testing.assert_series_equal(payroll_totals, expense_payroll, check_names=False)
    vendor_totals = data.vendors.groupby(["Month", "Department", "Expense_Category"])["Amount"].sum().sort_index()
    expense_non_payroll = data.expenses[data.expenses["Expense_Category"] != "Payroll"].groupby(["Month", "Department", "Expense_Category"])["Actual_Expense"].sum().sort_index()
    pd.testing.assert_series_equal(vendor_totals, expense_non_payroll, check_names=False)

    for path in [MONTHLY_XLSX, ANNUAL_XLSX, GOALS_PDF]:
        assert path.exists() and path.stat().st_size > 0, f"Missing or empty output: {path}"

    for workbook_path in [MONTHLY_XLSX, ANNUAL_XLSX]:
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
        expected_sheets = [
            "Executive_Summary", "Revenue", "Expenses", "Budget_vs_Actual",
            "Department_Summary", "Payroll", "Student_Payments", "Cash_Flow",
            "Scholarships", "Vendor_Payments", "Anomalies_Embedded",
        ]
        assert workbook.sheetnames == expected_sheets
        workbook.close()


def main() -> None:
    """Purpose: generate and verify all requested artifacts. Inputs: none. Outputs: files and console path confirmation."""

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    finance_data = generate_finance_data()
    create_workbook(finance_data, MONTHLY_XLSX, monthly=True)
    create_workbook(finance_data, ANNUAL_XLSX, monthly=False)
    create_goals_pdf(GOALS_PDF)
    validate_outputs(finance_data)

    print("Synthetic finance inputs generated and verified:")
    for path in [MONTHLY_XLSX, ANNUAL_XLSX, GOALS_PDF]:
        print(f" - {path.resolve()}")


if __name__ == "__main__":
    main()
