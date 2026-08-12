"""Tests for Phase 12A synthetic university financial history generation."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from finance_agent.calculations.calculation_loader import load_intermediate_model
from finance_agent.calculations.finance_engine import run_finance_calculations
from finance_agent.synthetic_history import SyntheticHistoryConfig, generate_synthetic_history, validate_generated_history
from finance_agent.understanding.intermediate import (
    build_financial_document_model,
    save_intermediate_outputs,
)


def _generate(tmp_path: Path, seed: int = 42, overwrite: bool = False):
    """Generate the default recovery scenario in a temporary folder.

    Inputs:
        tmp_path: Pytest temporary directory.
        seed: Deterministic random seed.
        overwrite: Whether generation may overwrite an existing folder.
    Outputs:
        Generated history object.
    Assumptions:
        Tests keep synthetic history isolated from repository fixtures.
    """

    return generate_synthetic_history(
        SyntheticHistoryConfig(output_directory=tmp_path, seed=seed, overwrite=overwrite)
    )


def _sheet_rows(path: Path, sheet: str) -> list[dict[str, object]]:
    """Read generated workbook rows from a named sheet.

    Inputs:
        path: Workbook path.
        sheet: Sheet name.
    Outputs:
        Row dictionaries using row 5 headers.
    Assumptions:
        Generated workbooks follow the current synthetic fixture layout.
    """

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    headers = [value for value in next(ws.iter_rows(min_row=5, max_row=5, values_only=True)) if value is not None]
    return [
        {header: values[idx] for idx, header in enumerate(headers)}
        for values in ws.iter_rows(min_row=6, values_only=True)
        if values[0] is not None
    ]


def test_generates_12_periods_with_valid_artifacts(tmp_path: Path) -> None:
    """Verify default generation creates 12 integrated workbooks and a manifest."""

    generated = _generate(tmp_path)

    assert len(generated.report_paths) == 12
    assert generated.manifest_path.exists()
    for report in generated.report_paths:
        wb = load_workbook(report, read_only=True, data_only=True)
        assert "Payroll" in wb.sheetnames
        assert "Financial_Position" in wb.sheetnames
        assert "Vendor_Payments" in wb.sheetnames
        assert "Anomalies_Embedded" in wb.sheetnames
        assert "Goals_Targets" in wb.sheetnames


def test_deterministic_seed_reproducibility(tmp_path: Path) -> None:
    """Verify identical seeds produce identical manifests and workbook data."""

    first = generate_synthetic_history(SyntheticHistoryConfig(output_directory=tmp_path / "a", seed=99))
    second = generate_synthetic_history(SyntheticHistoryConfig(output_directory=tmp_path / "b", seed=99))

    first_manifest = json.loads(first.manifest_path.read_text(encoding="utf-8"))
    second_manifest = json.loads(second.manifest_path.read_text(encoding="utf-8"))
    first_manifest["reports"] = []
    first_manifest["integrated_workbooks"] = []
    second_manifest["reports"] = []
    second_manifest["integrated_workbooks"] = []
    assert first_manifest == second_manifest
    assert _sheet_rows(first.report_paths[6], "Vendor_Payments") == _sheet_rows(second.report_paths[6], "Vendor_Payments")
    assert _sheet_rows(first.report_paths[6], "Payroll") == _sheet_rows(second.report_paths[6], "Payroll")


def test_monthly_and_annual_reconciliations(tmp_path: Path) -> None:
    """Verify generated rows reconcile monthly and annually."""

    generated = _generate(tmp_path)
    result = validate_generated_history(generated.root_directory)

    assert result.is_valid, result.errors
    manifest = generated.manifest
    assert result.reconciliations["annual"]["actual_revenue"] == manifest["annual_totals"]["actual_revenue"]
    assert result.reconciliations["annual"]["actual_expense"] == manifest["annual_totals"]["actual_expense"]


def test_coherent_cross_period_trends(tmp_path: Path) -> None:
    """Verify payroll pressure rises then stabilizes and collections recover."""

    generated = _generate(tmp_path)
    payroll_trend = generated.manifest["monthly_payroll_ratio_trend"]
    collection_trend = generated.manifest["collection_rate_trend"]

    assert payroll_trend["2026_06"] == max(payroll_trend.values())
    assert payroll_trend["2026_10"] < payroll_trend["2026_06"]
    assert collection_trend["2026_08"] > collection_trend["2026_07"]
    assert collection_trend["2026_12"] > collection_trend["2026_06"]


def test_expected_anomalies_appear_in_intended_months(tmp_path: Path) -> None:
    """Verify scenario anomalies appear only in expected manifest periods."""

    generated = _generate(tmp_path)
    manifest = generated.manifest

    assert manifest["health_sciences_overspending_periods"] == ["2026_04", "2026_05", "2026_06", "2026_07"]
    assert manifest["recurring_vendor_anomaly_periods"] == ["2026_07", "2026_08", "2026_09"]
    june_anomalies = _sheet_rows(generated.report_paths[5], "Anomalies_Embedded")
    assert any(row["Anomaly_Type"] == "negative_cash_flow" for row in june_anomalies)
    july_anomalies = _sheet_rows(generated.report_paths[6], "Anomalies_Embedded")
    assert any(row["Anomaly_Type"] == "recurring_vendor_duplicate" for row in july_anomalies)


def test_no_unintended_overwrites(tmp_path: Path) -> None:
    """Verify generator refuses to overwrite an existing scenario folder by default."""

    _generate(tmp_path)

    with pytest.raises(FileExistsError):
        _generate(tmp_path)


def test_manifest_matches_generated_data(tmp_path: Path) -> None:
    """Verify manifest totals and patterns match workbook contents."""

    generated = _generate(tmp_path)
    manifest = generated.manifest
    june_payroll_rows = _sheet_rows(generated.report_paths[5], "Payroll")
    june_revenue_rows = _sheet_rows(generated.report_paths[5], "Revenue")
    june_payroll = round(sum(float(row["Total_Payroll"]) for row in june_payroll_rows), 2)
    june_revenue = round(sum(float(row["Actual_Revenue"]) for row in june_revenue_rows), 2)

    assert round(june_payroll / june_revenue, 4) == manifest["monthly_payroll_ratio_trend"]["2026_06"]
    assert manifest["recommendation_milestone"]["period"] == "2026_05"


def test_financial_position_sheet_supports_health_ratios(tmp_path: Path) -> None:
    """Verify generated public fixtures contain deterministic ratio inputs."""

    generated = _generate(tmp_path)
    rows = _sheet_rows(generated.report_paths[8], "Financial_Position")

    assert rows
    row = rows[0]
    assert float(row["Current_Assets"]) > 0
    assert float(row["Current_Liabilities"]) > 0
    assert float(row["Cash_and_Equivalents"]) > 0
    assert float(row["Total_Assets"]) > float(row["Current_Assets"])
    assert "EBITDA" in row
    assert "Net_Income" in row


def test_representative_workbook_ratios_are_available_end_to_end(tmp_path: Path) -> None:
    """Verify the tracked representative workbook feeds all six health ratios."""

    report = Path(
        "data/synthetic_history/recovery_2026/reports/"
        "university_financial_report_2026_09.xlsx"
    )
    if not report.exists():
        pytest.skip("Tracked recovery history fixture is not available.")

    # This follows the real deterministic path used by the pipeline: understand
    # the workbook, persist normalized tables, reload them, then calculate KPIs.
    document_model = build_financial_document_model([report])
    position_tables = [
        table for table in document_model.tables if table.sheet == "Financial_Position"
    ]
    assert position_tables
    assert position_tables[0].detected_type == "Financial_Position"

    paths = save_intermediate_outputs(document_model, tmp_path / "intermediate")
    loaded_model = load_intermediate_model(paths["financial_document_model"])
    result = run_finance_calculations(
        loaded_model,
        source_workbook=str(report),
        report_period="2026_09",
    )
    ratios = {
        row["metric"]: row
        for row in result.kpi_summary.to_dict(orient="records")
        if row["metric"]
        in {
            "current_ratio",
            "cash_ratio",
            "total_debt_ratio",
            "ebitda_margin",
            "net_margin",
            "return_on_assets",
        }
    }

    expected = {
        "current_ratio": (1.4808053691275167, "Aceptable"),
        "cash_ratio": (0.7033557046979866, "Bueno"),
        "total_debt_ratio": (0.5209015095534677, "Aceptable"),
        "ebitda_margin": (0.2893226659434538, "Aceptable"),
        "net_margin": (-0.008239730000527343, "Crítico"),
        "return_on_assets": (-0.0018473556423519476, "Crítico"),
    }

    assert set(ratios) == set(expected)
    for metric, (value, classification) in expected.items():
        assert ratios[metric]["availability"] == "available"
        assert ratios[metric]["missing_inputs"] == []
        assert ratios[metric]["value"] == pytest.approx(value)
        assert ratios[metric]["classification"] == classification
