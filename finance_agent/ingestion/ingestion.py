"""Document ingestion and schema inspection utilities.

This module stops at extraction and structural inspection. It does not
normalize records, calculate KPIs, detect anomalies, or invoke an LLM.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from finance_agent.understanding.models import RawSheetData, RawWorkbookData


class WorkbookIngestionError(RuntimeError):
    """Raised when an Excel workbook or one of its sheets cannot be read."""


@dataclass(frozen=True)
class WorkbookIngestionResult:
    """Structured Excel ingestion result with metadata and loaded DataFrames."""

    workbook_path: str
    sheet_names: list[str]
    row_counts: dict[str, int]
    column_names: dict[str, list[str]]
    dataframes: dict[str, pd.DataFrame]

    def as_dict(self) -> dict[str, Any]:
        """Return all workbook fields as a dictionary.

        Inputs: this immutable result object.
        Outputs: metadata and loaded DataFrames in one dictionary.
        Assumptions: callers know DataFrames are not directly JSON serializable.
        """

        return {
            "workbook_path": self.workbook_path,
            "sheet_names": self.sheet_names,
            "row_counts": self.row_counts,
            "column_names": self.column_names,
            "dataframes": self.dataframes,
        }


def _validate_input_file(file_path: str | Path, expected_suffix: str) -> Path:
    """Validate and resolve an ingestion input path.

    Inputs: a file path and required lowercase suffix.
    Outputs: an absolute Path for an existing regular file.
    Assumptions: ingestion receives one explicit file, not a directory scan.
    """

    path = Path(file_path).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"Input file does not exist: {path}")
    if not path.is_file():
        raise FileNotFoundError(f"Input path is not a file: {path}")
    if path.suffix.lower() != expected_suffix:
        raise ValueError(
            f"Expected a {expected_suffix} file, received '{path.suffix or '<no suffix>'}': {path}"
        )
    return path.resolve()


def load_excel_workbook(
    workbook_path: str | Path,
    *,
    header_row: int | None = 0,
) -> WorkbookIngestionResult:
    """Load every sheet from an Excel workbook.

    Inputs: workbook path and optional zero-based pandas header row.
    Outputs: original sheet names, metadata, and DataFrames keyed by sheet name.
    Assumptions: every sheet uses the same header convention for this load.
    """

    path = _validate_input_file(workbook_path, ".xlsx")
    dataframes: dict[str, pd.DataFrame] = {}
    row_counts: dict[str, int] = {}
    column_names: dict[str, list[str]] = {}
    try:
        # The context manager releases the workbook handle immediately after all
        # sheets load, which matters for OneDrive-backed Windows workspaces.
        with pd.ExcelFile(path, engine="openpyxl") as excel_file:
            sheet_names = list(excel_file.sheet_names)
            for sheet_name in sheet_names:
                try:
                    dataframe = excel_file.parse(sheet_name=sheet_name, header=header_row)
                except Exception as error:
                    raise WorkbookIngestionError(
                        f"Unable to read sheet '{sheet_name}' from workbook '{path}': {error}"
                    ) from error

                # Messy workbook labels can be non-string values; converting
                # labels keeps inspection stable without altering cell data.
                dataframe.columns = [str(column) for column in dataframe.columns]
                dataframes[sheet_name] = dataframe
                row_counts[sheet_name] = len(dataframe.index)
                column_names[sheet_name] = list(dataframe.columns)
    except WorkbookIngestionError:
        raise
    except Exception as error:
        raise WorkbookIngestionError(f"Unable to open Excel workbook '{path}': {error}") from error

    return WorkbookIngestionResult(
        workbook_path=str(path),
        sheet_names=sheet_names,
        row_counts=row_counts,
        column_names=column_names,
        dataframes=dataframes,
    )


def load_raw_excel_workbook(workbook_path: str | Path) -> RawWorkbookData:
    """Load raw worksheet cells and layout evidence without assuming a header.

    Inputs: path to an Excel workbook.
    Outputs: original cell matrices, sheet order, and merged-cell ranges.
    Assumptions: formulas should be represented by their cached displayed values.
    """

    path = _validate_input_file(workbook_path, ".xlsx")
    try:
        # read_only=False is intentional: openpyxl exposes merged-cell geometry
        # only through a normal worksheet, and Step 2 needs that layout evidence.
        workbook = load_workbook(path, read_only=False, data_only=True)
    except Exception as error:
        raise WorkbookIngestionError(f"Unable to open raw workbook '{path}': {error}") from error

    sheets: dict[str, RawSheetData] = {}
    try:
        for worksheet in workbook.worksheets:
            # max_row/max_column can include styled empty cells. Trimming to the
            # last populated coordinate prevents formatting from becoming data.
            populated_coordinates = [
                (cell.row, cell.column)
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value is not None
            ]
            if populated_coordinates:
                max_row = max(row for row, _ in populated_coordinates)
                max_column = max(column for _, column in populated_coordinates)
                values = tuple(
                    tuple(
                        worksheet.cell(row=row_number, column=column_number).value
                        for column_number in range(1, max_column + 1)
                    )
                    for row_number in range(1, max_row + 1)
                )
            else:
                max_row = 0
                max_column = 0
                values = tuple()

            sheets[worksheet.title] = RawSheetData(
                sheet_name=worksheet.title,
                values=values,
                max_row=max_row,
                max_column=max_column,
                merged_ranges=tuple(
                    str(cell_range) for cell_range in worksheet.merged_cells.ranges
                ),
            )
    finally:
        # Releasing the handle matters when workbooks live in OneDrive folders.
        workbook.close()

    return RawWorkbookData(
        workbook_path=str(path),
        sheet_names=tuple(sheets),
        sheets=sheets,
    )


def _sample_rows_as_json_records(dataframe: pd.DataFrame, sample_size: int) -> list[dict[str, Any]]:
    """Convert sample rows to JSON-compatible records.

    Inputs: a DataFrame and maximum sample-row count.
    Outputs: records with ISO dates and null missing values.
    Assumptions: a compact sample is sufficient for later schema interpretation.
    """

    # pandas reliably converts Timestamp, date, NaN, and numpy scalar values.
    return json.loads(dataframe.head(sample_size).to_json(orient="records", date_format="iso"))


def inspect_sheet(
    sheet_name: str,
    dataframe: pd.DataFrame,
    *,
    sample_size: int = 5,
) -> dict[str, Any]:
    """Summarize the structure and data quality of one worksheet.

    Inputs: original sheet name, DataFrame, and sample-row count.
    Outputs: dimensions, labels, dtypes, sample rows, and missing counts.
    Assumptions: pandas dtypes are preliminary rather than a normalized schema.
    """

    if sample_size < 0:
        raise ValueError("sample_size must be zero or greater")

    return {
        "sheet_name": sheet_name,
        "row_count": int(len(dataframe.index)),
        "column_count": int(len(dataframe.columns)),
        "column_names": [str(column) for column in dataframe.columns],
        "inferred_data_types": {
            str(column): str(dtype) for column, dtype in dataframe.dtypes.items()
        },
        "sample_rows": _sample_rows_as_json_records(dataframe, sample_size),
        "missing_value_counts": {
            str(column): int(count) for column, count in dataframe.isna().sum().items()
        },
    }


def inspect_workbook(
    workbook: WorkbookIngestionResult,
    *,
    sample_size: int = 5,
) -> dict[str, Any]:
    """Create a JSON-compatible inspection summary for a workbook.

    Inputs: an ingested workbook and sample size per sheet.
    Outputs: workbook metadata and ordered sheet inspection summaries.
    Assumptions: source workbook sheet order remains meaningful.
    """

    return {
        "workbook_path": workbook.workbook_path,
        "sheet_count": len(workbook.sheet_names),
        "sheet_names": workbook.sheet_names,
        "sheets": [
            inspect_sheet(name, workbook.dataframes[name], sample_size=sample_size)
            for name in workbook.sheet_names
        ],
    }


def inspect_raw_workbook(
    workbook: RawWorkbookData,
    *,
    sample_size: int = 5,
) -> dict[str, Any]:
    """Inspect an already parsed raw workbook without reopening its source file.

    Inputs: canonical raw workbook cells and maximum sample rows.
    Outputs: the inspection shape consumed by preflight and persisted diagnostics.
    Assumptions: raw rows retain workbook order and may contain ragged/blank cells.
    """

    sheets: list[dict[str, Any]] = []
    for name in workbook.sheet_names:
        sheet = workbook.sheets[name]
        rows = list(sheet.values)
        header_index = min(4, max(0, len(rows) - 1)) if rows else 0
        header = rows[header_index] if rows else ()
        columns = [str(value) for value in header if value is not None]
        def json_cell(value: Any) -> Any:
            if isinstance(value, (date, datetime)):
                return value.isoformat()
            return value

        sheets.append(
            {
                "sheet_name": name,
                "row_count": max(0, len(rows) - header_index - 1),
                "column_count": sheet.max_column,
                "column_names": columns,
                "columns": columns,
                "sample_rows": [
                    [json_cell(value) for value in row]
                    for row in rows[header_index + 1 : header_index + 1 + sample_size]
                ],
            }
        )
    return {
        "workbook_path": workbook.workbook_path,
        "sheet_count": len(workbook.sheet_names),
        "sheet_names": list(workbook.sheet_names),
        "sheets": sheets,
    }
