"""Document ingestion and schema inspection utilities.

This module stops at extraction and structural inspection. It does not
normalize records, calculate KPIs, detect anomalies, or invoke an LLM.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader

from finance_agent.models import RawSheetData, RawWorkbookData


class WorkbookIngestionError(RuntimeError):
    """Raised when an Excel workbook or one of its sheets cannot be read."""


class PdfIngestionError(RuntimeError):
    """Raised when a goals PDF cannot be opened or its text cannot be extracted."""


@dataclass(frozen=True)
class WorkbookIngestionResult:
    """Structured Excel ingestion result with metadata and loaded DataFrames."""

    workbook_path: str
    sheet_names: list[str]
    row_counts: dict[str, int]
    column_names: dict[str, list[str]]
    dataframes: dict[str, pd.DataFrame]

    def as_dict(self) -> dict[str, Any]:
        """Return all workbook fields as a dictionary.

        Inputs: this immutable result object.
        Outputs: metadata and loaded DataFrames in one dictionary.
        Assumptions: callers know DataFrames are not directly JSON serializable.
        """

        return {
            "workbook_path": self.workbook_path,
            "sheet_names": self.sheet_names,
            "row_counts": self.row_counts,
            "column_names": self.column_names,
            "dataframes": self.dataframes,
        }


@dataclass(frozen=True)
class GoalsPdfResult:
    """Raw PDF text and basic file/document metadata."""

    pdf_path: str
    raw_text: str
    metadata: dict[str, Any]

    def as_dict(self) -> dict[str, Any]:
        """Return all PDF fields as a JSON-compatible dictionary.

        Inputs: this immutable result object.
        Outputs: PDF path, raw text, and metadata.
        Assumptions: metadata values were converted to JSON-safe types.
        """

        return {"pdf_path": self.pdf_path, "raw_text": self.raw_text, "metadata": self.metadata}


def _validate_input_file(file_path: str | Path, expected_suffix: str) -> Path:
    """Validate and resolve an ingestion input path.

    Inputs: a file path and required lowercase suffix.
    Outputs: an absolute Path for an existing regular file.
    Assumptions: ingestion receives one explicit file, not a directory scan.
    """

    path = Path(file_path).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"Input file does not exist: {path}")
    if not path.is_file():
        raise FileNotFoundError(f"Input path is not a file: {path}")
    if path.suffix.lower() != expected_suffix:
        raise ValueError(
            f"Expected a {expected_suffix} file, received '{path.suffix or '<no suffix>'}': {path}"
        )
    return path.resolve()


def load_excel_workbook(
    workbook_path: str | Path,
    *,
    header_row: int | None = 0,
) -> WorkbookIngestionResult:
    """Load every sheet from an Excel workbook.

    Inputs: workbook path and optional zero-based pandas header row.
    Outputs: original sheet names, metadata, and DataFrames keyed by sheet name.
    Assumptions: every sheet uses the same header convention for this load.
    """

    path = _validate_input_file(workbook_path, ".xlsx")
    dataframes: dict[str, pd.DataFrame] = {}
    row_counts: dict[str, int] = {}
    column_names: dict[str, list[str]] = {}
    try:
        # The context manager releases the workbook handle immediately after all
        # sheets load, which matters for OneDrive-backed Windows workspaces.
        with pd.ExcelFile(path, engine="openpyxl") as excel_file:
            sheet_names = list(excel_file.sheet_names)
            for sheet_name in sheet_names:
                try:
                    dataframe = excel_file.parse(sheet_name=sheet_name, header=header_row)
                except Exception as error:
                    raise WorkbookIngestionError(
                        f"Unable to read sheet '{sheet_name}' from workbook '{path}': {error}"
                    ) from error

                # Messy workbook labels can be non-string values; converting
                # labels keeps inspection stable without altering cell data.
                dataframe.columns = [str(column) for column in dataframe.columns]
                dataframes[sheet_name] = dataframe
                row_counts[sheet_name] = len(dataframe.index)
                column_names[sheet_name] = list(dataframe.columns)
    except WorkbookIngestionError:
        raise
    except Exception as error:
        raise WorkbookIngestionError(f"Unable to open Excel workbook '{path}': {error}") from error

    return WorkbookIngestionResult(
        workbook_path=str(path),
        sheet_names=sheet_names,
        row_counts=row_counts,
        column_names=column_names,
        dataframes=dataframes,
    )


def load_raw_excel_workbook(workbook_path: str | Path) -> RawWorkbookData:
    """Load raw worksheet cells and layout evidence without assuming a header.

    Inputs: path to an Excel workbook.
    Outputs: original cell matrices, sheet order, and merged-cell ranges.
    Assumptions: formulas should be represented by their cached displayed values.
    """

    path = _validate_input_file(workbook_path, ".xlsx")
    try:
        # read_only=False is intentional: openpyxl exposes merged-cell geometry
        # only through a normal worksheet, and Step 2 needs that layout evidence.
        workbook = load_workbook(path, read_only=False, data_only=True)
    except Exception as error:
        raise WorkbookIngestionError(f"Unable to open raw workbook '{path}': {error}") from error

    sheets: dict[str, RawSheetData] = {}
    try:
        for worksheet in workbook.worksheets:
            # max_row/max_column can include styled empty cells. Trimming to the
            # last populated coordinate prevents formatting from becoming data.
            populated_coordinates = [
                (cell.row, cell.column)
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value is not None
            ]
            if populated_coordinates:
                max_row = max(row for row, _ in populated_coordinates)
                max_column = max(column for _, column in populated_coordinates)
                values = tuple(
                    tuple(
                        worksheet.cell(row=row_number, column=column_number).value
                        for column_number in range(1, max_column + 1)
                    )
                    for row_number in range(1, max_row + 1)
                )
            else:
                max_row = 0
                max_column = 0
                values = tuple()

            sheets[worksheet.title] = RawSheetData(
                sheet_name=worksheet.title,
                values=values,
                max_row=max_row,
                max_column=max_column,
                merged_ranges=tuple(
                    str(cell_range) for cell_range in worksheet.merged_cells.ranges
                ),
            )
    finally:
        # Releasing the handle matters when workbooks live in OneDrive folders.
        workbook.close()

    return RawWorkbookData(
        workbook_path=str(path),
        sheet_names=tuple(sheets),
        sheets=sheets,
    )


def _sample_rows_as_json_records(dataframe: pd.DataFrame, sample_size: int) -> list[dict[str, Any]]:
    """Convert sample rows to JSON-compatible records.

    Inputs: a DataFrame and maximum sample-row count.
    Outputs: records with ISO dates and null missing values.
    Assumptions: a compact sample is sufficient for later schema interpretation.
    """

    # pandas reliably converts Timestamp, date, NaN, and numpy scalar values.
    return json.loads(dataframe.head(sample_size).to_json(orient="records", date_format="iso"))


def inspect_sheet(
    sheet_name: str,
    dataframe: pd.DataFrame,
    *,
    sample_size: int = 5,
) -> dict[str, Any]:
    """Summarize the structure and data quality of one worksheet.

    Inputs: original sheet name, DataFrame, and sample-row count.
    Outputs: dimensions, labels, dtypes, sample rows, and missing counts.
    Assumptions: pandas dtypes are preliminary rather than a normalized schema.
    """

    if sample_size < 0:
        raise ValueError("sample_size must be zero or greater")

    return {
        "sheet_name": sheet_name,
        "row_count": int(len(dataframe.index)),
        "column_count": int(len(dataframe.columns)),
        "column_names": [str(column) for column in dataframe.columns],
        "inferred_data_types": {
            str(column): str(dtype) for column, dtype in dataframe.dtypes.items()
        },
        "sample_rows": _sample_rows_as_json_records(dataframe, sample_size),
        "missing_value_counts": {
            str(column): int(count) for column, count in dataframe.isna().sum().items()
        },
    }


def inspect_workbook(
    workbook: WorkbookIngestionResult,
    *,
    sample_size: int = 5,
) -> dict[str, Any]:
    """Create a JSON-compatible inspection summary for a workbook.

    Inputs: an ingested workbook and sample size per sheet.
    Outputs: workbook metadata and ordered sheet inspection summaries.
    Assumptions: source workbook sheet order remains meaningful.
    """

    return {
        "workbook_path": workbook.workbook_path,
        "sheet_count": len(workbook.sheet_names),
        "sheet_names": workbook.sheet_names,
        "sheets": [
            inspect_sheet(name, workbook.dataframes[name], sample_size=sample_size)
            for name in workbook.sheet_names
        ],
    }


def extract_goals_pdf(pdf_path: str | Path) -> GoalsPdfResult:
    """Extract raw text and basic metadata from a goals PDF.

    Inputs: path to one PDF document.
    Outputs: concatenated page text and file/document metadata.
    Assumptions: advanced goal parsing and interpretation are deferred.
    """

    path = _validate_input_file(pdf_path, ".pdf")
    try:
        reader = PdfReader(path)
        page_text: list[str] = []
        for page_number, page in enumerate(reader.pages, start=1):
            try:
                # Separators retain page boundaries without imposing a goal schema.
                page_text.append((page.extract_text() or "").strip())
            except Exception as error:
                raise PdfIngestionError(
                    f"Unable to extract text from page {page_number} of '{path}': {error}"
                ) from error
    except PdfIngestionError:
        raise
    except Exception as error:
        raise PdfIngestionError(f"Unable to open PDF '{path}': {error}") from error

    document_metadata = reader.metadata or {}
    metadata = {
        "file_name": path.name,
        "file_size_bytes": path.stat().st_size,
        "page_count": len(reader.pages),
        "title": str(document_metadata.get("/Title") or ""),
        "author": str(document_metadata.get("/Author") or ""),
        "subject": str(document_metadata.get("/Subject") or ""),
        "creator": str(document_metadata.get("/Creator") or ""),
        "producer": str(document_metadata.get("/Producer") or ""),
        "creation_date": str(document_metadata.get("/CreationDate") or ""),
    }
    return GoalsPdfResult(
        pdf_path=str(path),
        raw_text="\n\n".join(text for text in page_text if text),
        metadata=metadata,
    )
