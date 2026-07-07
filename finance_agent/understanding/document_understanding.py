"""Deterministic worksheet layout analysis and logical-table detection."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl.utils.cell import range_boundaries

from finance_agent.understanding.models import (
    ContextRegion,
    DetectedRawTable,
    MergedTitleRegion,
    RawSheetData,
    RawWorkbookData,
    SheetUnderstanding,
)
from finance_agent.ingestion.schema import clean_column_name


HEADER_HINTS = {
    "actual",
    "allocated",
    "amount",
    "awarded",
    "budget",
    "cash",
    "category",
    "date",
    "department",
    "expense",
    "expenses",
    "fecha",
    "gasto",
    "gastos",
    "importe",
    "income",
    "ingreso",
    "ingresos",
    "invoice",
    "mes",
    "metric",
    "month",
    "monto",
    "nomina",
    "outstanding",
    "paid",
    "payroll",
    "payment",
    "period",
    "presupuesto",
    "program",
    "proveedor",
    "recipients",
    "remaining",
    "revenue",
    "salary",
    "scholarship",
    "status",
    "student",
    "total",
    "unidad",
    "valor",
    "variance",
    "vendor",
}

NOTE_HINTS = {
    "confidential",
    "currency",
    "fuente",
    "generated",
    "moneda",
    "nota",
    "note",
    "source",
    "synthetic",
}


@dataclass(frozen=True)
class HeaderCandidate:
    """Internal header candidate used while selecting logical tables."""

    row: int
    start_column: int
    end_column: int
    confidence: float


def _is_empty(value: Any) -> bool:
    """Determine whether a worksheet value is structurally empty.

    Inputs: one raw cell value.
    Outputs: True for None or whitespace-only strings.
    Assumptions: numeric zero and boolean False are meaningful values.
    """

    return value is None or (isinstance(value, str) and not value.strip())


def _cell_value(sheet: RawSheetData, row: int, column: int) -> Any:
    """Read a one-based cell from a raw worksheet matrix.

    Inputs: raw sheet and one-based row/column coordinates.
    Outputs: cell value or None when outside the populated matrix.
    Assumptions: out-of-range coordinates represent empty worksheet space.
    """

    if row < 1 or column < 1 or row > sheet.max_row or column > sheet.max_column:
        return None
    return sheet.values[row - 1][column - 1]


def _row_segments(sheet: RawSheetData, row: int, *, allowed_gap: int = 1) -> list[tuple[int, int]]:
    """Find horizontal groups of populated cells in one row.

    Inputs: raw sheet, row number, and tolerated internal blank-column gap.
    Outputs: one-based inclusive column segments.
    Assumptions: two or more blank columns usually separate side-by-side tables.
    """

    occupied = [
        column
        for column in range(1, sheet.max_column + 1)
        if not _is_empty(_cell_value(sheet, row, column))
    ]
    if not occupied:
        return []

    segments: list[tuple[int, int]] = []
    start = occupied[0]
    previous = occupied[0]
    for column in occupied[1:]:
        if column - previous > allowed_gap + 1:
            segments.append((start, previous))
            start = column
        previous = column
    segments.append((start, previous))
    return segments


def _value_kind(value: Any) -> str:
    """Classify a raw value broadly for header/data contrast.

    Inputs: one cell value.
    Outputs: text, numeric, date, boolean, or empty.
    Assumptions: broad kinds are sufficient for structural inference.
    """

    if _is_empty(value):
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return "date"
    if isinstance(value, (int, float)):
        return "numeric"
    return "text"


def _first_data_row(
    sheet: RawSheetData,
    header_row: int,
    start_column: int,
    end_column: int,
) -> int | None:
    """Find the first populated row directly below a potential header.

    Inputs: sheet and candidate header coordinates.
    Outputs: first populated row within the next three rows, or None.
    Assumptions: large gaps indicate that the candidate is not a table header.
    """

    for row in range(header_row + 1, min(sheet.max_row, header_row + 3) + 1):
        if any(
            not _is_empty(_cell_value(sheet, row, column))
            for column in range(start_column, end_column + 1)
        ):
            return row
    return None


def _score_header_candidate(
    sheet: RawSheetData,
    row: int,
    start_column: int,
    end_column: int,
    merged_rows: set[int],
) -> float:
    """Score one row segment as a potential header.

    Inputs: sheet, segment coordinates, and rows used by merged titles.
    Outputs: confidence between zero and one.
    Assumptions: headers are mostly short unique text with populated rows below.
    """

    if row in merged_rows:
        return 0.0

    values = [
        _cell_value(sheet, row, column)
        for column in range(start_column, end_column + 1)
    ]
    nonempty = [value for value in values if not _is_empty(value)]
    if len(nonempty) < 2:
        return 0.0

    text_values = [value for value in nonempty if isinstance(value, str)]
    text_ratio = len(text_values) / len(nonempty)
    if text_ratio < 0.60:
        return 0.0
    if text_values and sum(len(value.strip()) for value in text_values) / len(text_values) > 45:
        return 0.0

    data_row = _first_data_row(sheet, row, start_column, end_column)
    if data_row is None:
        return 0.0

    data_values = [
        _cell_value(sheet, data_row, column)
        for column in range(start_column, end_column + 1)
    ]
    populated_below = [value for value in data_values if not _is_empty(value)]
    below_density = len(populated_below) / max(1, end_column - start_column + 1)

    normalized_headers = [clean_column_name(value) for value in text_values]
    header_tokens = {
        token
        for header in normalized_headers
        for token in header.split("_")
        if token
    }
    keyword_score = min(1.0, len(header_tokens & HEADER_HINTS) / 2)
    uniqueness = len(set(normalized_headers)) / max(1, len(normalized_headers))
    type_contrast = 1.0 if any(
        _value_kind(value) in {"numeric", "date", "boolean"} for value in populated_below
    ) else 0.0

    confidence = (
        0.30 * text_ratio
        + 0.20 * below_density
        + 0.20 * keyword_score
        + 0.15 * type_contrast
        + 0.15 * uniqueness
    )
    return round(min(1.0, confidence), 4)


def _merged_title_regions(sheet: RawSheetData) -> list[MergedTitleRegion]:
    """Extract populated merged regions that behave like titles.

    Inputs: raw worksheet with merged-range evidence.
    Outputs: merged title/context regions.
    Assumptions: a merged region wider than one column is presentation context.
    """

    regions: list[MergedTitleRegion] = []
    for merged_range in sheet.merged_ranges:
        min_column, min_row, max_column, max_row = range_boundaries(merged_range)
        value = _cell_value(sheet, min_row, min_column)
        if max_column > min_column and not _is_empty(value):
            regions.append(
                MergedTitleRegion(
                    start_row=min_row,
                    end_row=max_row,
                    start_column=min_column,
                    end_column=max_column,
                    text=str(value).strip(),
                )
            )
    return regions


def _header_candidates(
    sheet: RawSheetData,
    merged_titles: list[MergedTitleRegion],
) -> list[HeaderCandidate]:
    """Discover plausible header segments throughout a worksheet.

    Inputs: raw sheet and merged title regions.
    Outputs: sorted header candidates above the structural threshold.
    Assumptions: confidence 0.62 balances flexible detection and false positives.
    """

    merged_rows = {
        row
        for region in merged_titles
        for row in range(region.start_row, region.end_row + 1)
    }
    candidates: list[HeaderCandidate] = []
    for row in range(1, sheet.max_row + 1):
        for start_column, end_column in _row_segments(sheet, row):
            confidence = _score_header_candidate(
                sheet,
                row,
                start_column,
                end_column,
                merged_rows,
            )
            if confidence >= 0.62:
                candidates.append(
                    HeaderCandidate(row, start_column, end_column, confidence)
                )
    return sorted(candidates, key=lambda item: (item.row, item.start_column))


def _candidate_overlap(left: HeaderCandidate, right: HeaderCandidate) -> float:
    """Measure horizontal overlap between two header candidates.

    Inputs: two candidate segments.
    Outputs: overlap fraction relative to the narrower segment.
    Assumptions: overlapping candidates likely describe successive vertical tables.
    """

    overlap = max(
        0,
        min(left.end_column, right.end_column)
        - max(left.start_column, right.start_column)
        + 1,
    )
    narrower = min(
        left.end_column - left.start_column + 1,
        right.end_column - right.start_column + 1,
    )
    return overlap / max(1, narrower)


def _looks_like_note_row(
    sheet: RawSheetData,
    row: int,
    start_column: int,
    end_column: int,
) -> bool:
    """Identify a sparse prose row that should not extend a table.

    Inputs: sheet, row, and active table columns.
    Outputs: True when the row resembles a note or footer.
    Assumptions: one long text cell after data is contextual prose.
    """

    values = [
        _cell_value(sheet, row, column)
        for column in range(start_column, end_column + 1)
        if not _is_empty(_cell_value(sheet, row, column))
    ]
    return len(values) == 1 and isinstance(values[0], str) and len(values[0].strip()) >= 45


def _table_end_row(
    sheet: RawSheetData,
    candidate: HeaderCandidate,
    candidates: list[HeaderCandidate],
) -> int:
    """Find the last data row belonging to a header candidate.

    Inputs: sheet, selected candidate, and all header candidates.
    Outputs: one-based inclusive end row, or the header row when no data exists.
    Assumptions: blank rows, new overlapping headers, and prose notes end a table.
    """

    last_data_row = candidate.row
    for row in range(candidate.row + 1, sheet.max_row + 1):
        values = [
            _cell_value(sheet, row, column)
            for column in range(candidate.start_column, candidate.end_column + 1)
        ]
        if all(_is_empty(value) for value in values):
            break

        later_headers = [
            other
            for other in candidates
            if other.row == row
            and _candidate_overlap(candidate, other) >= 0.5
            and other.confidence >= candidate.confidence
        ]
        if later_headers:
            break
        if last_data_row > candidate.row and _looks_like_note_row(
            sheet,
            row,
            candidate.start_column,
            candidate.end_column,
        ):
            break
        last_data_row = row
    return last_data_row


def _nearest_table_title(
    sheet: RawSheetData,
    header_row: int,
    start_column: int,
    end_column: int,
) -> str | None:
    """Find nearby contextual text immediately above a detected table.

    Inputs: sheet and table header coordinates.
    Outputs: nearest short title/context string, or None.
    Assumptions: titles usually appear within four rows above the header.
    """

    for row in range(header_row - 1, max(0, header_row - 5), -1):
        values = [
            _cell_value(sheet, row, column)
            for column in range(start_column, end_column + 1)
            if not _is_empty(_cell_value(sheet, row, column))
        ]
        if 1 <= len(values) <= 2:
            text = " | ".join(str(value).strip() for value in values)
            if len(text) <= 180:
                return text
    return None


def _unique_dataframe_columns(columns: list[str]) -> list[str]:
    """Create temporary unique DataFrame labels while retaining original labels separately.

    Inputs: original header labels.
    Outputs: non-empty unique labels suitable for pandas indexing.
    Assumptions: normalization later assigns the durable canonical names.
    """

    unique_columns: list[str] = []
    occurrences: dict[str, int] = {}
    for position, column in enumerate(columns, start=1):
        base = column.strip() or f"Unnamed Column {position}"
        occurrences[base] = occurrences.get(base, 0) + 1
        count = occurrences[base]
        unique_columns.append(base if count == 1 else f"{base} ({count})")
    return unique_columns


def _build_raw_table(
    sheet: RawSheetData,
    candidate: HeaderCandidate,
    end_row: int,
    table_index: int,
) -> DetectedRawTable:
    """Construct a positional DataFrame for one detected region.

    Inputs: sheet, header candidate, end row, and table index.
    Outputs: detected raw table with original coordinates and columns.
    Assumptions: the candidate header spans the intended table columns.
    """

    original_columns = [
        "" if _is_empty(_cell_value(sheet, candidate.row, column))
        else str(_cell_value(sheet, candidate.row, column)).strip()
        for column in range(candidate.start_column, candidate.end_column + 1)
    ]
    records = [
        [
            _cell_value(sheet, row, column)
            for column in range(candidate.start_column, candidate.end_column + 1)
        ]
        for row in range(candidate.row + 1, end_row + 1)
    ]
    dataframe = pd.DataFrame(
        records,
        columns=_unique_dataframe_columns(original_columns),
    ).dropna(how="all")
    return DetectedRawTable(
        table_index=table_index,
        sheet_name=sheet.sheet_name,
        header_row=candidate.row,
        start_row=candidate.row,
        end_row=end_row,
        start_column=candidate.start_column,
        end_column=candidate.end_column,
        header_confidence=candidate.confidence,
        title=_nearest_table_title(
            sheet,
            candidate.row,
            candidate.start_column,
            candidate.end_column,
        ),
        original_columns=original_columns,
        dataframe=dataframe.reset_index(drop=True),
    )


def _context_regions(
    sheet: RawSheetData,
    tables: list[DetectedRawTable],
) -> list[ContextRegion]:
    """Identify non-table populated row regions as title, note, or footer context.

    Inputs: raw sheet and selected tables.
    Outputs: contextual row ranges with combined text.
    Assumptions: row-level context is sufficient; cell-level table geometry is preserved separately.
    """

    covered_rows = {
        row
        for table in tables
        for row in range(table.start_row, table.end_row + 1)
    }
    uncovered_rows = [
        row
        for row in range(1, sheet.max_row + 1)
        if row not in covered_rows
        and any(
            not _is_empty(_cell_value(sheet, row, column))
            for column in range(1, sheet.max_column + 1)
        )
    ]
    if not uncovered_rows:
        return []

    groups: list[list[int]] = [[uncovered_rows[0]]]
    for row in uncovered_rows[1:]:
        if row == groups[-1][-1] + 1:
            groups[-1].append(row)
        else:
            groups.append([row])

    first_table_row = min((table.start_row for table in tables), default=sheet.max_row + 1)
    last_table_row = max((table.end_row for table in tables), default=0)
    regions: list[ContextRegion] = []
    for group in groups:
        text_parts = [
            str(_cell_value(sheet, row, column)).strip()
            for row in group
            for column in range(1, sheet.max_column + 1)
            if not _is_empty(_cell_value(sheet, row, column))
        ]
        text = " | ".join(dict.fromkeys(text_parts))
        normalized_tokens = set(clean_column_name(text).split("_"))
        if group[-1] < first_table_row:
            region_type = "title_or_context"
        elif group[0] > last_table_row:
            region_type = "notes_or_footer"
        elif normalized_tokens & NOTE_HINTS:
            region_type = "note"
        else:
            region_type = "context"
        regions.append(
            ContextRegion(
                start_row=group[0],
                end_row=group[-1],
                region_type=region_type,
                text=text,
            )
        )
    return regions


def understand_sheet(sheet: RawSheetData) -> SheetUnderstanding:
    """Detect structural regions and logical tables in one worksheet.

    Inputs: raw worksheet values and layout evidence.
    Outputs: table regions, separators, merged titles, and contextual regions.
    Assumptions: deterministic confidence may leave ambiguous regions as context.
    """

    merged_titles = _merged_title_regions(sheet)
    candidates = _header_candidates(sheet, merged_titles)
    tables: list[DetectedRawTable] = []
    selected_regions: list[tuple[int, int, int, int]] = []

    for candidate in candidates:
        end_row = _table_end_row(sheet, candidate, candidates)
        if end_row <= candidate.row:
            continue

        # A candidate inside an already selected table is usually a text-heavy
        # data row. Side-by-side candidates remain eligible when columns differ.
        overlaps_existing = any(
            candidate.row <= selected_end_row
            and end_row >= selected_start_row
            and candidate.start_column <= selected_end_column
            and candidate.end_column >= selected_start_column
            for (
                selected_start_row,
                selected_end_row,
                selected_start_column,
                selected_end_column,
            ) in selected_regions
        )
        if overlaps_existing:
            continue

        table = _build_raw_table(
            sheet,
            candidate,
            end_row,
            table_index=len(tables) + 1,
        )
        if table.dataframe.empty:
            continue
        tables.append(table)
        selected_regions.append(
            (
                table.start_row,
                table.end_row,
                table.start_column,
                table.end_column,
            )
        )

    empty_rows = [
        row
        for row in range(1, sheet.max_row + 1)
        if all(
            _is_empty(_cell_value(sheet, row, column))
            for column in range(1, sheet.max_column + 1)
        )
    ]
    return SheetUnderstanding(
        sheet_name=sheet.sheet_name,
        max_row=sheet.max_row,
        max_column=sheet.max_column,
        empty_separator_rows=empty_rows,
        merged_title_regions=merged_titles,
        context_regions=_context_regions(sheet, tables),
        tables=tables,
    )


def understand_workbook(workbook: RawWorkbookData) -> list[SheetUnderstanding]:
    """Analyze every worksheet in a raw workbook.

    Inputs: raw workbook ingestion result.
    Outputs: structural analysis in original sheet order.
    Assumptions: each sheet is understood independently before model assembly.
    """

    return [
        understand_sheet(workbook.sheets[sheet_name])
        for sheet_name in workbook.sheet_names
    ]
