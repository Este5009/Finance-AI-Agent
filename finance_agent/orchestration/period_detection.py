"""Deterministic period detection for generic pipeline inputs."""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path
from typing import Any

from finance_agent.ingestion.ingestion import extract_goals_pdf
from finance_agent.orchestration.pipeline_models import DetectedPeriod, PipelineInputModel


MONTH_ALIASES: dict[str, int] = {
    "january": 1,
    "jan": 1,
    "enero": 1,
    "february": 2,
    "feb": 2,
    "febrero": 2,
    "march": 3,
    "mar": 3,
    "marzo": 3,
    "april": 4,
    "apr": 4,
    "abril": 4,
    "may": 5,
    "mayo": 5,
    "june": 6,
    "jun": 6,
    "junio": 6,
    "july": 7,
    "jul": 7,
    "julio": 7,
    "august": 8,
    "aug": 8,
    "agosto": 8,
    "september": 9,
    "sep": 9,
    "sept": 9,
    "septiembre": 9,
    "october": 10,
    "oct": 10,
    "octubre": 10,
    "november": 11,
    "nov": 11,
    "noviembre": 11,
    "december": 12,
    "dec": 12,
    "diciembre": 12,
}

LOW_CONFIDENCE_PERIOD = DetectedPeriod(
    period_type="unknown",
    label="Unknown period",
    confidence=0.0,
    evidence=("No reliable period evidence found.",),
)


def _normalize_text(value: str) -> str:
    """Normalize text for deterministic period matching.

    Inputs: arbitrary text.
    Outputs: lowercase text with separators simplified.
    Assumptions: matching should be conservative and language-light.
    """

    return re.sub(r"[_\-.]+", " ", value.casefold())


def _years(text: str) -> list[int]:
    """Extract plausible reporting years from text.

    Inputs: normalized or raw text.
    Outputs: ordered list of 2000..2099 years.
    Assumptions: this project focuses on modern financial reports.
    """

    return [int(match) for match in re.findall(r"\b(20\d{2})\b", text)]


def _month_from_text(text: str) -> int | None:
    """Find a month name or number in text.

    Inputs: normalized text.
    Outputs: month number or None.
    Assumptions: named months are stronger evidence than bare numbers.
    """

    for token, month in MONTH_ALIASES.items():
        if re.search(rf"\b{re.escape(token)}\b", text):
            return month
    numeric = re.search(r"\b(?:month|mes)\s*0?([1-9]|1[0-2])\b", text)
    return int(numeric.group(1)) if numeric else None


def _detect_from_text(text: str, *, source_label: str, base_confidence: float) -> DetectedPeriod | None:
    """Detect a period from one text source.

    Inputs: text, source label, and base confidence for that source.
    Outputs: detected period or None.
    Assumptions: filename evidence is usually stronger than sparse workbook text.
    """

    raw = text.casefold()
    normalized = _normalize_text(text)
    years = _years(normalized)
    year = years[0] if years else None
    evidence: list[str] = []

    range_match = re.search(
        r"\b(20\d{2})[-/](0?[1-9]|1[0-2])[-/](0?[1-9]|[12]\d|3[01])\s*(?:to|a|-)\s*"
        r"(20\d{2})[-/](0?[1-9]|1[0-2])[-/](0?[1-9]|[12]\d|3[01])\b",
        raw,
    )
    if range_match:
        start = f"{range_match.group(1)}-{int(range_match.group(2)):02d}-{int(range_match.group(3)):02d}"
        end = f"{range_match.group(4)}-{int(range_match.group(5)):02d}-{int(range_match.group(6)):02d}"
        return DetectedPeriod(
            period_type="custom",
            label=f"{start} to {end}",
            confidence=min(1.0, base_confidence + 0.12),
            start_date=start,
            end_date=end,
            evidence=(f"{source_label}: explicit date range",),
        )

    quarter_match = re.search(r"\b(?:q|quarter|trimestre)\s*([1-4])\b", normalized)
    if quarter_match and year:
        quarter = int(quarter_match.group(1))
        return DetectedPeriod(
            period_type="quarterly",
            label=f"Q{quarter} {year}",
            confidence=min(1.0, base_confidence + 0.08),
            year=year,
            quarter=quarter,
            evidence=(f"{source_label}: quarter and year",),
        )

    semester_match = re.search(r"\b(?:s|semester|semestre)\s*([12])\b", normalized)
    if semester_match and year:
        semester = int(semester_match.group(1))
        return DetectedPeriod(
            period_type="semester",
            label=f"S{semester} {year}",
            confidence=min(1.0, base_confidence + 0.08),
            year=year,
            semester=semester,
            evidence=(f"{source_label}: semester and year",),
        )

    if year and re.search(r"\b(?:annual|anual|yearly|fiscal year|año)\b", normalized):
        return DetectedPeriod(
            period_type="annual",
            label=str(year),
            confidence=min(1.0, base_confidence + 0.1),
            year=year,
            evidence=(f"{source_label}: annual keyword and year",),
        )

    month = _month_from_text(normalized)
    if month and year:
        return DetectedPeriod(
            period_type="monthly",
            label=f"{year}-{month:02d}",
            confidence=min(1.0, base_confidence + 0.1),
            year=year,
            month=month,
            evidence=(f"{source_label}: month and year",),
        )
    if year:
        evidence.append(f"{source_label}: year only")
        return DetectedPeriod(
            period_type="unknown",
            label=str(year),
            confidence=max(0.25, base_confidence - 0.25),
            year=year,
            evidence=tuple(evidence),
        )
    return None


def _sample_workbook_text(path: Path, *, max_cells: int = 500) -> str:
    """Collect bounded workbook text for period detection.

    Inputs: workbook path and maximum cells to inspect.
    Outputs: concatenated cell and sheet-name text.
    Assumptions: this is metadata sampling, not report ingestion or calculation.
    """

    if path.suffix.casefold() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return ""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    values: list[str] = list(workbook.sheetnames)
    seen = 0
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        values.append(str(cell.value))
                        seen += 1
                    if seen >= max_cells:
                        return " ".join(values)
    finally:
        workbook.close()
    return " ".join(values)


def _sample_goals_text(path: Path, *, max_chars: int = 4000) -> str:
    """Extract bounded goals text for period detection.

    Inputs: goals document path.
    Outputs: text snippet.
    Assumptions: PDF and plain text are supported now; DOCX can be added later.
    """

    suffix = path.suffix.casefold()
    try:
        if suffix == ".pdf":
            return extract_goals_pdf(path).raw_text[:max_chars]
        if suffix in {".txt", ".md"}:
            return path.read_text(encoding="utf-8", errors="ignore")[:max_chars]
    except Exception:
        return ""
    return ""


def _choose_detection(candidates: list[DetectedPeriod]) -> DetectedPeriod:
    """Choose the strongest non-conflicting period candidate.

    Inputs: candidate detections from filename, workbook, and goals text.
    Outputs: one DetectedPeriod.
    Assumptions: conflicts reduce confidence instead of guessing.
    """

    if not candidates:
        return LOW_CONFIDENCE_PERIOD
    keyed = Counter((candidate.period_type, candidate.label) for candidate in candidates)
    best = max(candidates, key=lambda item: (keyed[(item.period_type, item.label)], item.confidence))
    conflicts = {
        (candidate.period_type, candidate.label)
        for candidate in candidates
        if candidate.period_type != "unknown"
    }
    evidence = tuple(
        dict.fromkeys(
            evidence for candidate in candidates for evidence in candidate.evidence
        )
    )
    confidence = best.confidence
    if len(conflicts) > 1:
        if keyed[(best.period_type, best.label)] >= 2:
            # Multiple independent report-side signals should be allowed to
            # outweigh a broader goals document, while still recording caution.
            confidence = max(0.65, min(confidence, best.confidence - 0.05))
            evidence = (*evidence, "Conflicting lower-weight period candidate found.")
        else:
            confidence = min(confidence, 0.55)
            evidence = (*evidence, "Conflicting period candidates found.")
    return DetectedPeriod(
        period_type=best.period_type if confidence >= 0.65 else "unknown",
        label=best.label,
        confidence=confidence,
        evidence=evidence,
        year=best.year,
        month=best.month,
        quarter=best.quarter,
        semester=best.semester,
        start_date=best.start_date,
        end_date=best.end_date,
    )


def detect_period(
    financial_report_path: str | Path,
    goals_document_path: str | Path | None = None,
) -> DetectedPeriod:
    """Infer the reporting period for one financial report.

    Inputs: financial report path and optional goals document path.
    Outputs: DetectedPeriod with confidence and evidence.
    Assumptions: low-confidence output requires a user override before execution.
    """

    report = Path(financial_report_path)
    candidates: list[DetectedPeriod] = []
    for source_label, text, confidence in (
        ("filename", report.stem, 0.78),
        ("workbook", _sample_workbook_text(report), 0.58),
    ):
        if text and (detected := _detect_from_text(text, source_label=source_label, base_confidence=confidence)):
            candidates.append(detected)
    if goals_document_path:
        goals = Path(goals_document_path)
        goals_text = f"{goals.stem} {_sample_goals_text(goals)}"
        if detected := _detect_from_text(goals_text, source_label="goals", base_confidence=0.54):
            candidates.append(detected)
    return _choose_detection(candidates)


def build_pipeline_input_model(
    *,
    financial_report_path: str | Path,
    goals_document_path: str | Path,
    period_override: str | None = None,
    report_language: str = "es",
) -> PipelineInputModel:
    """Build the generic one-report pipeline input model.

    Inputs: report path, goals path, optional period override, and language.
    Outputs: PipelineInputModel for orchestrator or future UI use.
    Assumptions: language affects user-facing report text, not internal field names.
    """

    detected = detect_period(financial_report_path, goals_document_path)
    period_type = detected.period_type if not period_override else _detect_override_type(period_override)
    return PipelineInputModel(
        financial_report_path=Path(financial_report_path).resolve(),
        goals_document_path=Path(goals_document_path).resolve(),
        detected_period=detected,
        period_type=period_type,
        period_override=period_override,
        report_language=report_language or "es",
    )


def _detect_override_type(period_override: str) -> str:
    """Infer a broad period type from a user override label.

    Inputs: user-supplied period override.
    Outputs: period type string.
    Assumptions: override text is authoritative even if only loosely parsed.
    """

    detected = _detect_from_text(period_override, source_label="override", base_confidence=0.9)
    return detected.period_type if detected else "custom"
