"""Generator for coherent synthetic university financial histories."""

from __future__ import annotations

import json
import random
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from finance_agent.synthetic_history.models import GeneratedHistory, MonthlyFinancialData, SyntheticHistoryConfig
from finance_agent.synthetic_history.scenarios import MONTH_NAMES, MONTH_NAMES_ES, get_scenario_points


SHEET_ORDER = [
    "Executive_Summary",
    "Revenue",
    "Expenses",
    "Budget_vs_Actual",
    "Department_Summary",
    "Payroll",
    "Student_Payments",
    "Cash_Flow",
    "Scholarships",
    "Vendor_Payments",
    "Anomalies_Embedded",
]


def generate_synthetic_history(config: SyntheticHistoryConfig | None = None) -> GeneratedHistory:
    """Generate a deterministic multi-period synthetic university history.

    Inputs:
        config: Optional generation configuration. Defaults produce the 2026
            recovery scenario under ``data/synthetic_history/recovery_2026``.
    Outputs:
        Generated artifact paths and the manifest dictionary.
    Assumptions:
        Existing scenario folders are protected unless ``overwrite=True``.
    """

    active_config = config or SyntheticHistoryConfig()
    scenario_root = active_config.output_directory / active_config.scenario_slug
    reports_dir = scenario_root / "reports"
    goals_dir = scenario_root / "goals"

    _prepare_output_directory(scenario_root, active_config.overwrite)
    reports_dir.mkdir(parents=True, exist_ok=True)
    goals_dir.mkdir(parents=True, exist_ok=True)

    rng = random.Random(active_config.seed)
    scenario_points = get_scenario_points(active_config.scenario)
    generated_months: list[MonthlyFinancialData] = []
    report_paths: list[Path] = []
    goals_paths: list[Path] = []

    beginning_cash = 2_500_000.0
    for point in scenario_points:
        monthly_data = _build_monthly_financial_data(active_config, point, rng, beginning_cash)
        generated_months.append(monthly_data)
        beginning_cash = monthly_data.totals["ending_cash"]

        report_path = reports_dir / f"university_financial_report_{active_config.year}_{point.month:02d}.xlsx"
        goals_path = goals_dir / f"financial_goals_{active_config.year}_{point.month:02d}.pdf"
        _write_workbook(report_path, active_config.year, point.month, monthly_data)
        _write_goals_pdf(goals_path, active_config.year, point.month, point, monthly_data)
        report_paths.append(report_path)
        goals_paths.append(goals_path)

    manifest = _build_manifest(active_config, generated_months, report_paths, goals_paths)
    manifest_path = scenario_root / "scenario_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    return GeneratedHistory(
        root_directory=scenario_root,
        report_paths=report_paths,
        goals_paths=goals_paths,
        manifest_path=manifest_path,
        manifest=manifest,
    )


def _prepare_output_directory(path: Path, overwrite: bool) -> None:
    """Create or safely reset the scenario output directory.

    Inputs:
        path: Scenario directory that will receive generated artifacts.
        overwrite: Whether to delete an existing scenario directory first.
    Outputs:
        None.
    Assumptions:
        The caller passes the intended scenario directory, never a shared root.
    """

    if path.exists() and any(path.iterdir()):
        if not overwrite:
            raise FileExistsError(f"Synthetic history output already exists: {path}")
        # Destructive cleanup is restricted to the computed scenario folder.
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def _build_monthly_financial_data(
    config: SyntheticHistoryConfig,
    point: Any,
    rng: random.Random,
    beginning_cash: float,
) -> MonthlyFinancialData:
    """Build reconciled rows for a single month.

    Inputs:
        config: Active generator configuration.
        point: Monthly scenario controls.
        rng: Deterministic random generator for small non-semantic variation.
        beginning_cash: Cash balance carried from the previous month.
    Outputs:
        Monthly financial rows keyed by workbook sheet name.
    Assumptions:
        Revenue, expense, budget, department, and cash-flow rows reconcile.
    """

    period_date = date(config.year, point.month, 1)
    period_slug = f"{config.year}_{point.month:02d}"
    month_name = MONTH_NAMES[point.month]
    month_name_es = MONTH_NAMES_ES[point.month]
    departments = list(config.departments)

    base_monthly_revenue = 2_100_000.0
    revenue_budget_total = base_monthly_revenue * (1 + 0.004 * (point.month - 1))
    actual_revenue_total = revenue_budget_total * point.revenue_factor
    payroll_total = actual_revenue_total * point.payroll_ratio
    expense_total = actual_revenue_total - (point.net_cash_flow * 0.55)
    expense_budget_total = revenue_budget_total * 0.93
    scholarships_total = actual_revenue_total * 0.055
    capital_outflows = max(90_000.0, 170_000.0 - 5_000 * point.month)
    actual_cash_inflows = actual_revenue_total * point.collection_rate + 210_000.0
    actual_operating_outflows = actual_cash_inflows - scholarships_total - capital_outflows - point.net_cash_flow
    ending_cash = beginning_cash + point.net_cash_flow

    dept_weights = _department_weights(departments)
    revenue_rows = _build_revenue_rows(period_date, month_name, departments, dept_weights, revenue_budget_total, actual_revenue_total)
    payroll_rows = _build_payroll_rows(
        period_date,
        month_name,
        departments,
        dept_weights,
        payroll_total,
        point.health_sciences_overtime_factor,
    )
    expense_rows = _build_expense_rows(period_date, month_name, departments, dept_weights, expense_budget_total, expense_total, payroll_rows)
    budget_rows = _build_budget_rows(period_date, month_name, departments, revenue_rows, expense_rows)
    department_rows = _build_department_summary_rows(budget_rows)
    payments_rows = _build_student_payment_rows(period_slug, period_date, month_name, departments, dept_weights, actual_revenue_total, point.collection_rate)
    scholarships_rows = _build_scholarship_rows(period_date, month_name, departments, dept_weights, scholarships_total)
    vendor_rows = _build_vendor_payment_rows(period_slug, period_date, month_name, departments, rng, point.vendor_anomaly)
    cash_rows = [
        {
            "Period": datetime(config.year, point.month, 1),
            "Month": month_name,
            "Beginning_Cash": round(beginning_cash, 2),
            "Budget_Cash_Inflows": round(revenue_budget_total * 0.93 + 210_000.0, 2),
            "Actual_Cash_Inflows": round(actual_cash_inflows, 2),
            "Budget_Operating_Outflows": round(expense_budget_total * 0.78, 2),
            "Actual_Operating_Outflows": round(actual_operating_outflows, 2),
            "Budget_Scholarships": round(revenue_budget_total * 0.052, 2),
            "Actual_Scholarships": round(scholarships_total, 2),
            "Budget_Capital_Outflows": round(capital_outflows * 0.95, 2),
            "Actual_Capital_Outflows": round(capital_outflows, 2),
            "Budget_Net_Cash_Flow": round(80_000.0 + 15_000.0 * min(point.month, 6), 2),
            "Actual_Net_Cash_Flow": round(point.net_cash_flow, 2),
            "Budget_Ending_Cash": round(beginning_cash + 80_000.0 + 15_000.0 * min(point.month, 6), 2),
            "Actual_Ending_Cash": round(ending_cash, 2),
            "Ending_Cash_Variance": round(ending_cash - (beginning_cash + 80_000.0 + 15_000.0 * min(point.month, 6)), 2),
        }
    ]
    anomalies_rows = _build_anomaly_rows(period_slug, point, department_rows, payroll_rows, vendor_rows, point.net_cash_flow)
    executive_rows = _build_executive_summary_rows(
        actual_revenue_total,
        revenue_budget_total,
        expense_total,
        expense_budget_total,
        payroll_total,
        point.collection_rate,
        point.net_cash_flow,
        anomalies_rows,
    )

    totals = {
        "actual_revenue": round(actual_revenue_total, 2),
        "budget_revenue": round(revenue_budget_total, 2),
        "actual_expense": round(expense_total, 2),
        "budget_expense": round(expense_budget_total, 2),
        "payroll_total": round(payroll_total, 2),
        "payroll_ratio": round(payroll_total / actual_revenue_total, 4),
        "collection_rate": round(point.collection_rate, 4),
        "net_cash_flow": round(point.net_cash_flow, 2),
        "ending_cash": round(ending_cash, 2),
    }
    rows_by_sheet = {
        "Executive_Summary": executive_rows,
        "Revenue": revenue_rows,
        "Expenses": expense_rows,
        "Budget_vs_Actual": budget_rows,
        "Department_Summary": department_rows,
        "Payroll": payroll_rows,
        "Student_Payments": payments_rows,
        "Cash_Flow": cash_rows,
        "Scholarships": scholarships_rows,
        "Vendor_Payments": vendor_rows,
        "Anomalies_Embedded": anomalies_rows,
    }
    return MonthlyFinancialData(period_slug, month_name, rows_by_sheet, totals, [row["Anomaly_ID"] for row in anomalies_rows])


def _department_weights(departments: list[str]) -> dict[str, float]:
    """Return deterministic allocation weights for departments.

    Inputs:
        departments: Department names from configuration.
    Outputs:
        Weight mapping that sums to one.
    Assumptions:
        Known departments receive domain-shaped weights; unknown custom
        departments share the remaining weight evenly.
    """

    preferred = {
        "Engineering": 0.22,
        "Business": 0.20,
        "Health Sciences": 0.26,
        "Arts & Humanities": 0.14,
        "Student Services": 0.10,
        "Administration": 0.08,
    }
    weights = {dept: preferred.get(dept, 1.0) for dept in departments}
    total = sum(weights.values())
    return {dept: value / total for dept, value in weights.items()}


def _build_revenue_rows(
    period_date: date,
    month_name: str,
    departments: list[str],
    weights: dict[str, float],
    budget_total: float,
    actual_total: float,
) -> list[dict[str, Any]]:
    """Build revenue rows that reconcile to monthly totals.

    Inputs:
        period_date: Month start date.
        month_name: English month name used by current synthetic workbooks.
        departments: Department list.
        weights: Department allocation weights.
        budget_total: Total budget revenue.
        actual_total: Total actual revenue.
    Outputs:
        Revenue rows split by department and category.
    Assumptions:
        Tuition is the largest revenue category for every department.
    """

    categories = {"Tuition": 0.78, "Grants": 0.12, "Auxiliary": 0.07, "Donations": 0.03}
    rows: list[dict[str, Any]] = []
    for dept in departments:
        for category, category_weight in categories.items():
            budget = budget_total * weights[dept] * category_weight
            actual = actual_total * weights[dept] * category_weight
            rows.append(_variance_row(period_date, month_name, dept, "Revenue_Category", category, budget, actual, "Revenue"))
    _adjust_rows_to_total(rows, "Actual_Revenue", actual_total)
    _adjust_rows_to_total(rows, "Budget_Revenue", budget_total)
    return rows


def _build_payroll_rows(
    period_date: date,
    month_name: str,
    departments: list[str],
    weights: dict[str, float],
    payroll_total: float,
    health_sciences_overtime_factor: float,
) -> list[dict[str, Any]]:
    """Build department payroll rows including overtime evidence.

    Inputs:
        period_date: Month start date.
        month_name: Month name.
        departments: Department list.
        weights: Department allocation weights.
        payroll_total: Target payroll total.
        health_sciences_overtime_factor: Scenario overtime pressure multiplier.
    Outputs:
        Payroll rows with headcount, salary, benefits, overtime, budget, and variance.
    Assumptions:
        Health Sciences overtime is the primary payroll anomaly driver.
    """

    rows: list[dict[str, Any]] = []
    overtime_multipliers = {dept: 1.0 for dept in departments}
    if "Health Sciences" in overtime_multipliers:
        overtime_multipliers["Health Sciences"] = health_sciences_overtime_factor
    overtime_weight_total = sum(weights[d] * (0.08 * overtime_multipliers[d]) for d in departments)
    base_benefit_weight_total = sum(weights[d] * 0.92 for d in departments)

    for dept in departments:
        dept_total = payroll_total * (weights[dept] * (0.92 + 0.08 * overtime_multipliers[dept])) / (base_benefit_weight_total + overtime_weight_total)
        overtime = dept_total * min(0.22, 0.08 * overtime_multipliers[dept])
        benefits = dept_total * 0.19
        base_salary = dept_total - benefits - overtime
        budget = dept_total / (1.16 if dept == "Health Sciences" and health_sciences_overtime_factor > 1.4 else 1.04)
        variance = dept_total - budget
        rows.append(
            {
                "Period": datetime(period_date.year, period_date.month, 1),
                "Month": month_name,
                "Department": dept,
                "Headcount_FTE": round(18 + weights[dept] * 120, 1),
                "Base_Salary": round(base_salary, 2),
                "Benefits": round(benefits, 2),
                "Overtime": round(overtime, 2),
                "Total_Payroll": round(dept_total, 2),
                "Payroll_Budget": round(budget, 2),
                "Variance": round(variance, 2),
                "Variance_Pct": round(variance / budget if budget else 0.0, 4),
            }
        )
    _adjust_rows_to_total(rows, "Total_Payroll", payroll_total)
    return rows


def _build_expense_rows(
    period_date: date,
    month_name: str,
    departments: list[str],
    weights: dict[str, float],
    budget_total: float,
    actual_total: float,
    payroll_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build expense rows that include payroll and operating costs.

    Inputs:
        period_date: Month start date.
        month_name: Month name.
        departments: Department list.
        weights: Department weights.
        budget_total: Total budget expense.
        actual_total: Total actual expense.
        payroll_rows: Payroll rows used to keep payroll embedded in expenses.
    Outputs:
        Expense rows split by department and category.
    Assumptions:
        The Expenses sheet is the all-in operating expense view.
    """

    payroll_by_dept = {row["Department"]: float(row["Total_Payroll"]) for row in payroll_rows}
    non_payroll_total = max(actual_total - sum(payroll_by_dept.values()), actual_total * 0.32)
    non_payroll_categories = {"Supplies": 0.38, "Services": 0.34, "Facilities": 0.18, "Technology": 0.10}
    rows: list[dict[str, Any]] = []
    for dept in departments:
        payroll_actual = payroll_by_dept[dept]
        payroll_budget = payroll_actual / 1.04
        rows.append(_variance_row(period_date, month_name, dept, "Expense_Category", "Payroll", payroll_budget, payroll_actual, "Expense"))
        for category, category_weight in non_payroll_categories.items():
            budget = (budget_total - sum(payroll_by_dept.values()) / 1.04) * weights[dept] * category_weight
            actual = non_payroll_total * weights[dept] * category_weight
            rows.append(_variance_row(period_date, month_name, dept, "Expense_Category", category, budget, actual, "Expense"))
    _adjust_rows_to_total(rows, "Actual_Expense", actual_total)
    _adjust_rows_to_total(rows, "Budget_Expense", budget_total)
    return rows


def _variance_row(
    period_date: date,
    month_name: str,
    department: str,
    category_field: str,
    category: str,
    budget: float,
    actual: float,
    row_type: str,
) -> dict[str, Any]:
    """Build one budget-vs-actual row for revenue or expense sheets.

    Inputs:
        period_date: Month start date.
        month_name: Month name.
        department: Department name.
        category_field: Sheet-specific category column.
        category: Category value.
        budget: Budget amount.
        actual: Actual amount.
        row_type: ``Revenue`` or ``Expense``.
    Outputs:
        A row with the naming convention expected by existing synthetic inputs.
    Assumptions:
        Variance is actual minus budget.
    """

    row = {"Period": datetime(period_date.year, period_date.month, 1), "Month": month_name, "Department": department, category_field: category}
    if row_type == "Revenue":
        row.update(
            {
                "Budget_Revenue": round(budget, 2),
                "Actual_Revenue": round(actual, 2),
                "Variance": round(actual - budget, 2),
                "Variance_Pct": round((actual - budget) / budget if budget else 0.0, 4),
            }
        )
    else:
        row.update(
            {
                "Budget_Expense": round(budget, 2),
                "Actual_Expense": round(actual, 2),
                "Variance": round(actual - budget, 2),
                "Variance_Pct": round((actual - budget) / budget if budget else 0.0, 4),
            }
        )
    return row


def _build_budget_rows(
    period_date: date,
    month_name: str,
    departments: list[str],
    revenue_rows: list[dict[str, Any]],
    expense_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build department-level budget versus actual rows.

    Inputs:
        period_date: Month start date.
        month_name: Month name.
        departments: Department list.
        revenue_rows: Revenue detail rows.
        expense_rows: Expense detail rows.
    Outputs:
        Department-level budget rows.
    Assumptions:
        Revenue and expense detail rows use current synthetic column names.
    """

    rows: list[dict[str, Any]] = []
    for dept in departments:
        budget_revenue = sum(float(row["Budget_Revenue"]) for row in revenue_rows if row["Department"] == dept)
        actual_revenue = sum(float(row["Actual_Revenue"]) for row in revenue_rows if row["Department"] == dept)
        budget_expense = sum(float(row["Budget_Expense"]) for row in expense_rows if row["Department"] == dept)
        actual_expense = sum(float(row["Actual_Expense"]) for row in expense_rows if row["Department"] == dept)
        revenue_variance = actual_revenue - budget_revenue
        expense_variance = actual_expense - budget_expense
        rows.append(
            {
                "Period": datetime(period_date.year, period_date.month, 1),
                "Month": month_name,
                "Department": dept,
                "Budget_Revenue": round(budget_revenue, 2),
                "Actual_Revenue": round(actual_revenue, 2),
                "Budget_Expense": round(budget_expense, 2),
                "Actual_Expense": round(actual_expense, 2),
                "Revenue_Variance": round(revenue_variance, 2),
                "Revenue_Variance_Pct": round(revenue_variance / budget_revenue if budget_revenue else 0.0, 4),
                "Expense_Variance": round(expense_variance, 2),
                "Expense_Variance_Pct": round(expense_variance / budget_expense if budget_expense else 0.0, 4),
                "Net_Budget": round(budget_revenue - budget_expense, 2),
                "Net_Actual": round(actual_revenue - actual_expense, 2),
            }
        )
    return rows


def _build_department_summary_rows(budget_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Build department summary rows from budget rows.

    Inputs:
        budget_rows: Department budget-vs-actual rows.
    Outputs:
        Department summary rows with a readable status.
    Assumptions:
        Expense variance percentage above 8% is an attention flag.
    """

    rows: list[dict[str, Any]] = []
    for row in budget_rows:
        variance_pct = float(row["Expense_Variance_Pct"])
        if row["Department"] == "Health Sciences" and variance_pct > 0.08:
            status = "Over budget - overtime pressure"
        elif variance_pct > 0.05:
            status = "Over budget"
        elif float(row["Net_Actual"]) < 0:
            status = "Negative contribution"
        else:
            status = "On track"
        rows.append(
            {
                "Department": row["Department"],
                "Budget_Revenue": row["Budget_Revenue"],
                "Actual_Revenue": row["Actual_Revenue"],
                "Budget_Expense": row["Budget_Expense"],
                "Actual_Expense": row["Actual_Expense"],
                "Net_Contribution": row["Net_Actual"],
                "Expense_Variance": row["Expense_Variance"],
                "Expense_Variance_Pct": row["Expense_Variance_Pct"],
                "Status": status,
            }
        )
    return rows


def _build_student_payment_rows(
    period_slug: str,
    period_date: date,
    month_name: str,
    departments: list[str],
    weights: dict[str, float],
    revenue_total: float,
    collection_rate: float,
) -> list[dict[str, Any]]:
    """Build student payment collection rows.

    Inputs:
        period_slug: Canonical period slug.
        period_date: Month start date.
        month_name: Month name.
        departments: Department list.
        weights: Department weights.
        revenue_total: Actual revenue basis.
        collection_rate: Scenario collection rate.
    Outputs:
        Student payment rows with due, paid, outstanding, and overdue fields.
    Assumptions:
        Student payment rows are aggregated examples, not individual student PII.
    """

    rows: list[dict[str, Any]] = []
    cohorts = ["Freshman", "Sophomore", "Junior", "Senior"]
    invoice_counter = 1
    for dept in departments:
        due_total = revenue_total * weights[dept] * 0.72
        for cohort in cohorts:
            amount_due = due_total / len(cohorts)
            amount_paid = amount_due * collection_rate
            outstanding = amount_due - amount_paid
            rows.append(
                {
                    "Invoice_ID": f"INV-{period_slug}-{invoice_counter:03d}",
                    "Student_ID": f"{dept[:3].upper()}-{cohort[:2].upper()}-{invoice_counter:03d}",
                    "Department": dept,
                    "Billing_Period": month_name,
                    "Due_Date": datetime(period_date.year, period_date.month, 20),
                    "Payment_Date": datetime(period_date.year, period_date.month, 25) if collection_rate >= 0.9 else None,
                    "Amount_Due": round(amount_due, 2),
                    "Amount_Paid": round(amount_paid, 2),
                    "Outstanding": round(outstanding, 2),
                    "Status": "Collected" if collection_rate >= 0.92 else "Partially collected",
                    "Days_Overdue": 0 if collection_rate >= 0.9 else 18,
                }
            )
            invoice_counter += 1
    return rows


def _build_scholarship_rows(
    period_date: date,
    month_name: str,
    departments: list[str],
    weights: dict[str, float],
    scholarships_total: float,
) -> list[dict[str, Any]]:
    """Build scholarship utilization rows.

    Inputs:
        period_date: Month start date.
        month_name: Month name.
        departments: Department list.
        weights: Department weights.
        scholarships_total: Total awarded scholarships.
    Outputs:
        Scholarship rows compatible with current synthetic workbook structure.
    Assumptions:
        Scholarships are included for continuity with existing synthetic fixtures.
    """

    rows: list[dict[str, Any]] = []
    for dept in departments:
        allocated = scholarships_total * weights[dept] * 1.08
        awarded = scholarships_total * weights[dept]
        rows.append(
            {
                "Period": datetime(period_date.year, period_date.month, 1),
                "Month": month_name,
                "Department": dept,
                "Scholarship_Type": "Need-based",
                "Allocated": round(allocated, 2),
                "Awarded": round(awarded, 2),
                "Remaining": round(allocated - awarded, 2),
                "Recipients": max(8, round(weights[dept] * 220)),
                "Utilization_Pct": round(awarded / allocated if allocated else 0.0, 4),
            }
        )
    return rows


def _build_vendor_payment_rows(
    period_slug: str,
    period_date: date,
    month_name: str,
    departments: list[str],
    rng: random.Random,
    vendor_anomaly: bool,
) -> list[dict[str, Any]]:
    """Build vendor payment rows and optional recurring anomaly evidence.

    Inputs:
        period_slug: Canonical period slug.
        period_date: Month start date.
        month_name: Month name.
        departments: Department list.
        rng: Deterministic random source.
        vendor_anomaly: Whether to encode duplicate/high-value vendor payments.
    Outputs:
        Vendor payment rows.
    Assumptions:
        ``MedSupply Co`` is the recurring high-risk vendor in the recovery scenario.
    """

    vendors = ["CampusWorks", "LabSource", "FacilitiesPlus", "CloudEdu", "LibraryHub", "MedSupply Co"]
    rows: list[dict[str, Any]] = []
    for idx, dept in enumerate(departments, start=1):
        vendor = vendors[idx % len(vendors)]
        amount = 18_000 + rng.randint(0, 7_500) + idx * 1_250
        rows.append(
            {
                "Payment_ID": f"PAY-{period_slug}-{idx:03d}",
                "Payment_Date": datetime(period_date.year, period_date.month, min(25, 5 + idx * 3)),
                "Month": month_name,
                "Department": dept,
                "Vendor": vendor,
                "Invoice_Number": f"{vendor[:3].upper()}-{period_slug}-{idx:03d}",
                "Expense_Category": "Services" if vendor != "MedSupply Co" else "Medical Supplies",
                "Amount": round(amount, 2),
                "Payment_Method": "ACH",
                "Approval_Status": "Approved",
                "Potential_Duplicate": False,
                "High_Value_Flag": amount > 30_000,
            }
        )
    if vendor_anomaly:
        for duplicate_idx in range(2):
            rows.append(
                {
                    "Payment_ID": f"PAY-{period_slug}-MED-{duplicate_idx + 1}",
                    "Payment_Date": datetime(period_date.year, period_date.month, 18 + duplicate_idx),
                    "Month": month_name,
                    "Department": "Health Sciences",
                    "Vendor": "MedSupply Co",
                    "Invoice_Number": f"MED-{period_slug}-777",
                    "Expense_Category": "Medical Supplies",
                    "Amount": 74_500.0,
                    "Payment_Method": "ACH",
                    "Approval_Status": "Approved - expedited",
                    "Potential_Duplicate": duplicate_idx == 1,
                    "High_Value_Flag": True,
                }
            )
    return rows


def _build_anomaly_rows(
    period_slug: str,
    point: Any,
    department_rows: list[dict[str, Any]],
    payroll_rows: list[dict[str, Any]],
    vendor_rows: list[dict[str, Any]],
    net_cash_flow: float,
) -> list[dict[str, Any]]:
    """Build known anomaly-supporting rows for the scenario manifest.

    Inputs:
        period_slug: Canonical period slug.
        point: Monthly scenario controls.
        department_rows: Department summary rows.
        payroll_rows: Payroll rows.
        vendor_rows: Vendor payment rows.
        net_cash_flow: Monthly actual net cash flow.
    Outputs:
        Embedded anomaly evidence rows.
    Assumptions:
        These rows are expected test evidence, not results from anomaly detection.
    """

    rows: list[dict[str, Any]] = []
    hs_summary = next((row for row in department_rows if row["Department"] == "Health Sciences"), None)
    hs_payroll = next((row for row in payroll_rows if row["Department"] == "Health Sciences"), None)
    if hs_summary and hs_payroll and point.health_sciences_overtime_factor >= 1.6:
        rows.append(
            {
                "Anomaly_ID": f"ANOM-{period_slug}-HS-OT",
                "Detected_Period": period_slug,
                "Department": "Health Sciences",
                "Anomaly_Type": "payroll_overtime_overspend",
                "Severity": "High" if point.month < 6 else "Critical",
                "Description": "Health Sciences overtime remains above the reduction threshold.",
                "Observed_Value": round(float(hs_payroll["Overtime"]), 2),
                "Threshold": round(float(hs_payroll["Payroll_Budget"]) * 0.08, 2),
                "Source_Sheet": "Payroll",
            }
        )
    if net_cash_flow <= -300_000:
        rows.append(
            {
                "Anomaly_ID": f"ANOM-{period_slug}-CF",
                "Detected_Period": period_slug,
                "Department": "University",
                "Anomaly_Type": "negative_cash_flow",
                "Severity": "Critical" if net_cash_flow <= -600_000 else "High",
                "Description": "Net cash flow is materially negative for the month.",
                "Observed_Value": round(net_cash_flow, 2),
                "Threshold": 0.0,
                "Source_Sheet": "Cash_Flow",
            }
        )
    if any(row["Potential_Duplicate"] for row in vendor_rows):
        rows.append(
            {
                "Anomaly_ID": f"ANOM-{period_slug}-VENDOR",
                "Detected_Period": period_slug,
                "Department": "Health Sciences",
                "Anomaly_Type": "recurring_vendor_duplicate",
                "Severity": "High",
                "Description": "Recurring MedSupply Co payment shares an invoice number and high-value flag.",
                "Observed_Value": 149_000.0,
                "Threshold": 50_000.0,
                "Source_Sheet": "Vendor_Payments",
            }
        )
    return rows


def _build_executive_summary_rows(
    actual_revenue: float,
    budget_revenue: float,
    actual_expense: float,
    budget_expense: float,
    payroll_total: float,
    collection_rate: float,
    net_cash_flow: float,
    anomaly_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build the monthly executive summary sheet rows.

    Inputs:
        actual_revenue: Monthly actual revenue.
        budget_revenue: Monthly budget revenue.
        actual_expense: Monthly actual expense.
        budget_expense: Monthly budget expense.
        payroll_total: Monthly payroll total.
        collection_rate: Student collection rate.
        net_cash_flow: Actual net cash flow.
        anomaly_rows: Embedded anomaly evidence rows.
    Outputs:
        Executive summary metric rows.
    Assumptions:
        Summary rows expose enough context for ingestion tests without analysis.
    """

    payroll_ratio = payroll_total / actual_revenue if actual_revenue else 0.0
    return [
        _summary_row("Actual Revenue", actual_revenue, budget_revenue, "Currency", "On target" if actual_revenue >= budget_revenue * 0.98 else "Below target"),
        _summary_row("Operating Expenses", actual_expense, budget_expense, "Currency", "Over budget" if actual_expense > budget_expense else "On budget"),
        _summary_row("Net Operating Result", actual_revenue - actual_expense, budget_revenue - budget_expense, "Currency", "Negative" if actual_revenue < actual_expense else "Positive"),
        _summary_row("Payroll Ratio", payroll_ratio, 0.42, "Percent", "High" if payroll_ratio > 0.48 else "Healthy"),
        _summary_row("Collection Rate", collection_rate, 0.93, "Percent", "Below target" if collection_rate < 0.90 else "On target"),
        _summary_row("Net Cash Flow", net_cash_flow, 0.0, "Currency", "Weak" if net_cash_flow < 0 else "Positive"),
        _summary_row("Embedded Anomaly Count", len(anomaly_rows), 0, "Count", "Review" if anomaly_rows else "Clear"),
    ]


def _summary_row(metric: str, actual: float, goal: float, fmt: str, status: str) -> dict[str, Any]:
    """Build one executive summary row.

    Inputs:
        metric: Metric name.
        actual: Actual value.
        goal: Goal or budget value.
        fmt: Display format hint.
        status: Human-readable status.
    Outputs:
        A summary row.
    Assumptions:
        Variance is actual minus goal.
    """

    return {
        "Metric": metric,
        "Actual": round(actual, 4) if fmt == "Percent" else round(actual, 2),
        "Goal_or_Budget": round(goal, 4) if fmt == "Percent" else round(goal, 2),
        "Variance": round(actual - goal, 4) if fmt == "Percent" else round(actual - goal, 2),
        "Format": fmt,
        "Status": status,
    }


def _adjust_rows_to_total(rows: list[dict[str, Any]], key: str, expected_total: float) -> None:
    """Adjust the final row to remove rounding drift.

    Inputs:
        rows: Rows containing numeric values.
        key: Numeric key to reconcile.
        expected_total: Target sum.
    Outputs:
        None; mutates the final row's value and variance fields when present.
    Assumptions:
        Rounding drift is small and the final row is safe to adjust.
    """

    if not rows:
        return
    current_total = sum(float(row.get(key, 0.0)) for row in rows)
    drift = round(expected_total - current_total, 2)
    if abs(drift) >= 0.01:
        rows[-1][key] = round(float(rows[-1][key]) + drift, 2)
        if "Budget_Revenue" in rows[-1] and "Actual_Revenue" in rows[-1]:
            rows[-1]["Variance"] = round(float(rows[-1]["Actual_Revenue"]) - float(rows[-1]["Budget_Revenue"]), 2)
            rows[-1]["Variance_Pct"] = round(rows[-1]["Variance"] / float(rows[-1]["Budget_Revenue"]) if rows[-1]["Budget_Revenue"] else 0.0, 4)
        if "Budget_Expense" in rows[-1] and "Actual_Expense" in rows[-1]:
            rows[-1]["Variance"] = round(float(rows[-1]["Actual_Expense"]) - float(rows[-1]["Budget_Expense"]), 2)
            rows[-1]["Variance_Pct"] = round(rows[-1]["Variance"] / float(rows[-1]["Budget_Expense"]) if rows[-1]["Budget_Expense"] else 0.0, 4)


def _write_workbook(path: Path, year: int, month: int, monthly_data: MonthlyFinancialData) -> None:
    """Write one monthly financial report workbook.

    Inputs:
        path: Target workbook path.
        year: Report year.
        month: Report month.
        monthly_data: Generated rows and totals.
    Outputs:
        An ``.xlsx`` workbook on disk.
    Assumptions:
        Header row 5 mirrors the current synthetic fixture format.
    """

    wb = Workbook()
    # Stable workbook metadata keeps generated artifacts as reproducible as the
    # XLSX container format allows and avoids timestamp-only differences.
    stable_timestamp = datetime(year, 1, 1, 0, 0, 0)
    wb.properties.creator = "Finance AI Agent Synthetic History"
    wb.properties.created = stable_timestamp
    wb.properties.modified = stable_timestamp
    default = wb.active
    wb.remove(default)
    for sheet_name in SHEET_ORDER:
        ws = wb.create_sheet(sheet_name)
        rows = monthly_data.rows_by_sheet[sheet_name]
        title = f"{sheet_name.replace('_', ' ')} - {MONTH_NAMES[month]} {year}"
        _write_sheet(ws, title, rows)
    wb.save(path)


def _write_sheet(ws: Any, title: str, rows: list[dict[str, Any]]) -> None:
    """Write one professional workbook sheet.

    Inputs:
        ws: OpenPyXL worksheet.
        title: Sheet title.
        rows: Row dictionaries.
    Outputs:
        Worksheet content and basic formatting.
    Assumptions:
        The first row defines the ordered set of headers.
    """

    ws["A1"] = title
    ws["A2"] = "Small/medium university | Synthetic recovery history"
    ws["A3"] = "Currency: USD | Synthetic data | Generated deterministically"
    for cell in ("A1", "A2", "A3"):
        ws[cell].font = Font(bold=cell == "A1")
    headers = list(rows[0].keys()) if rows else ["No_Data"]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_idx, row in enumerate(rows, start=6):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header))
    ws.freeze_panes = "A6"
    for col_idx, header in enumerate(headers, start=1):
        width = max(12, min(28, len(str(header)) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_goals_pdf(path: Path, year: int, month: int, point: Any, monthly_data: MonthlyFinancialData) -> None:
    """Write a Spanish monthly goals PDF.

    Inputs:
        path: Target PDF path.
        year: Report year.
        month: Report month.
        point: Monthly scenario controls.
        monthly_data: Generated financial totals.
    Outputs:
        A valid PDF goals document.
    Assumptions:
        Goals are concise user-facing context, not source-of-truth calculations.
    """

    doc = SimpleDocTemplate(str(path), pagesize=letter, title=f"Metas financieras {year}-{month:02d}")
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph(f"Metas financieras universitarias - {MONTH_NAMES_ES[month]} {year}", styles["Title"]),
        Spacer(1, 12),
        Paragraph(point.narrative_es, styles["BodyText"]),
        Spacer(1, 12),
    ]
    goal_rows = [
        ["Meta", "Objetivo mensual", "Progreso esperado"],
        ["Margen operativo", "Mantener resultado operativo positivo", _goal_status(monthly_data.totals["actual_revenue"] - monthly_data.totals["actual_expense"], 0)],
        ["Ratio de nómina", "Mantener nómina bajo 42% de ingresos", f"{monthly_data.totals['payroll_ratio']:.1%}"],
        ["Cobranza estudiantil", "Recuperar al menos 93% de saldos facturados", f"{monthly_data.totals['collection_rate']:.1%}"],
        ["Flujo de caja", "Evitar flujo de caja mensual negativo", _goal_status(monthly_data.totals["net_cash_flow"], 0)],
    ]
    table = Table(goal_rows, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.extend([table, Spacer(1, 12)])
    if point.recommendation_milestone or point.policy_action_es:
        story.append(Paragraph(f"Hito de gestión: {point.policy_action_es}", styles["BodyText"]))
    else:
        story.append(Paragraph("Hito de gestión: monitoreo mensual de ingresos, gastos, nómina y cobranza.", styles["BodyText"]))
    doc.build(story)


def _goal_status(actual: float, target: float) -> str:
    """Format a concise Spanish goal status.

    Inputs:
        actual: Actual value.
        target: Threshold value.
    Outputs:
        User-facing Spanish status string.
    Assumptions:
        Higher than target is favorable for these goal rows.
    """

    status = "cumple" if actual >= target else "requiere atención"
    return f"{actual:,.0f} ({status})"


def _build_manifest(
    config: SyntheticHistoryConfig,
    months: list[MonthlyFinancialData],
    report_paths: list[Path],
    goals_paths: list[Path],
) -> dict[str, Any]:
    """Build the scenario manifest for later assertions.

    Inputs:
        config: Active generator configuration.
        months: Generated monthly data.
        report_paths: Workbook artifact paths.
        goals_paths: Goals PDF artifact paths.
    Outputs:
        JSON-serializable manifest dictionary.
    Assumptions:
        Manifest values are derived from generated rows, not independently copied.
    """

    hs_overspend_periods = [
        month.period_slug
        for month in months
        if any(row["Anomaly_Type"] == "payroll_overtime_overspend" for row in month.rows_by_sheet["Anomalies_Embedded"])
    ]
    vendor_anomaly_periods = [
        month.period_slug
        for month in months
        if any(row["Anomaly_Type"] == "recurring_vendor_duplicate" for row in month.rows_by_sheet["Anomalies_Embedded"])
    ]
    cash_flow_recovery_periods = [month.period_slug for month in months if month.totals["net_cash_flow"] > 0 and int(month.period_slug[-2:]) >= 9]
    annual_totals = {
        "actual_revenue": round(sum(month.totals["actual_revenue"] for month in months), 2),
        "actual_expense": round(sum(month.totals["actual_expense"] for month in months), 2),
        "payroll_total": round(sum(month.totals["payroll_total"] for month in months), 2),
        "net_cash_flow": round(sum(month.totals["net_cash_flow"] for month in months), 2),
    }
    return {
        "scenario": config.scenario,
        "year": config.year,
        "seed": config.seed,
        "period_type": "monthly",
        "departments": list(config.departments),
        "reports": [str(path.as_posix()) for path in report_paths],
        "goals": [str(path.as_posix()) for path in goals_paths],
        "monthly_payroll_ratio_trend": {month.period_slug: month.totals["payroll_ratio"] for month in months},
        "collection_rate_trend": {month.period_slug: month.totals["collection_rate"] for month in months},
        "health_sciences_overspending_periods": hs_overspend_periods,
        "recurring_vendor_anomaly_periods": vendor_anomaly_periods,
        "recommendation_milestone": {"period": f"{config.year}_05", "topic": "health_sciences_overtime_reduction"},
        "cash_flow_recovery_periods": cash_flow_recovery_periods,
        "expected_goal_progress": {
            "payroll_ratio_peak": f"{config.year}_06",
            "collection_campaign_start": f"{config.year}_08",
            "hiring_freeze_start": f"{config.year}_09",
            "cash_flow_positive_from": f"{config.year}_09",
        },
        "monthly_totals": {month.period_slug: month.totals for month in months},
        "annual_totals": annual_totals,
        "expected_timeline_es": {
            month.period_slug: get_scenario_points(config.scenario)[idx].narrative_es for idx, month in enumerate(months)
        },
    }
