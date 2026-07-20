"""Validation helpers for generated synthetic financial histories."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from finance_agent.synthetic_history.models import SyntheticHistoryValidationResult


def validate_generated_history(root_directory: str | Path) -> SyntheticHistoryValidationResult:
    """Validate generated synthetic history artifacts and reconciliations.

    Inputs:
        root_directory: Scenario root containing reports, goals, and manifest.
    Outputs:
        Structured validation result with errors, warnings, and reconciliations.
    Assumptions:
        Workbooks use the Phase 12A header-on-row-5 synthetic format.
    """

    root = Path(root_directory)
    errors: list[str] = []
    warnings: list[str] = []
    reconciliations: dict[str, Any] = {"monthly": {}, "annual": {}}

    manifest_path = root / "scenario_manifest.json"
    if not manifest_path.exists():
        return SyntheticHistoryValidationResult(False, [f"Missing manifest: {manifest_path}"], warnings, reconciliations)

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    report_paths = [Path(path) for path in manifest.get("reports", [])]
    goals_paths = [Path(path) for path in manifest.get("goals", [])]
    if len(report_paths) != 12:
        errors.append(f"Expected 12 reports, found {len(report_paths)}")
    if len(goals_paths) != 12:
        errors.append(f"Expected 12 goals documents, found {len(goals_paths)}")

    annual_revenue = 0.0
    annual_expense = 0.0
    annual_payroll = 0.0
    annual_cash_flow = 0.0
    workbook_periods: list[str] = []
    for report_path in report_paths:
        if not report_path.exists():
            errors.append(f"Missing report workbook: {report_path}")
            continue
        workbook_result = _validate_workbook(report_path)
        errors.extend(workbook_result["errors"])
        period = workbook_result["period"]
        if period:
            workbook_periods.append(period)
            reconciliations["monthly"][period] = workbook_result["totals"]
            annual_revenue += workbook_result["totals"]["actual_revenue"]
            annual_expense += workbook_result["totals"]["actual_expense"]
            annual_payroll += workbook_result["totals"]["payroll_total"]
            annual_cash_flow += workbook_result["totals"]["net_cash_flow"]

    for goals_path in goals_paths:
        if not goals_path.exists():
            errors.append(f"Missing goals PDF: {goals_path}")
            continue
        try:
            PdfReader(str(goals_path))
        except Exception as exc:  # pragma: no cover - defensive PDF parser boundary
            errors.append(f"Invalid goals PDF {goals_path}: {exc}")

    manifest_totals = manifest.get("annual_totals", {})
    computed_totals = {
        "actual_revenue": round(annual_revenue, 2),
        "actual_expense": round(annual_expense, 2),
        "payroll_total": round(annual_payroll, 2),
        "net_cash_flow": round(annual_cash_flow, 2),
    }
    reconciliations["annual"] = computed_totals
    for key, value in computed_totals.items():
        if round(float(manifest_totals.get(key, 0.0)), 2) != value:
            errors.append(f"Annual {key} does not match manifest: {value} != {manifest_totals.get(key)}")

    _validate_manifest_patterns(manifest, workbook_periods, errors)
    return SyntheticHistoryValidationResult(not errors, errors, warnings, reconciliations)


def _validate_workbook(path: Path) -> dict[str, Any]:
    """Validate one generated workbook and return reconciled totals.

    Inputs:
        path: Workbook path.
    Outputs:
        Dict containing errors, period, and totals.
    Assumptions:
        Required sheets contain row 5 headers.
    """

    errors: list[str] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    required = {
        "Revenue",
        "Expenses",
        "Budget_vs_Actual",
        "Department_Summary",
        "Payroll",
        "Student_Payments",
        "Cash_Flow",
        "Vendor_Payments",
        "Anomalies_Embedded",
    }
    missing_sheets = required - set(wb.sheetnames)
    if missing_sheets:
        errors.append(f"{path.name} missing sheets: {sorted(missing_sheets)}")

    revenue_rows = _read_sheet_rows(wb, "Revenue")
    expense_rows = _read_sheet_rows(wb, "Expenses")
    payroll_rows = _read_sheet_rows(wb, "Payroll")
    budget_rows = _read_sheet_rows(wb, "Budget_vs_Actual")
    cash_rows = _read_sheet_rows(wb, "Cash_Flow")
    period = _period_from_path(path)

    actual_revenue = round(sum(float(row.get("Actual_Revenue") or 0.0) for row in revenue_rows), 2)
    actual_expense = round(sum(float(row.get("Actual_Expense") or 0.0) for row in expense_rows), 2)
    payroll_total = round(sum(float(row.get("Total_Payroll") or 0.0) for row in payroll_rows), 2)
    budget_revenue = round(sum(float(row.get("Budget_Revenue") or 0.0) for row in revenue_rows), 2)
    budget_expense = round(sum(float(row.get("Budget_Expense") or 0.0) for row in expense_rows), 2)
    cash_flow = round(float(cash_rows[0].get("Actual_Net_Cash_Flow") or 0.0), 2) if cash_rows else 0.0

    budget_revenue_total = round(sum(float(row.get("Actual_Revenue") or 0.0) for row in budget_rows), 2)
    budget_expense_total = round(sum(float(row.get("Actual_Expense") or 0.0) for row in budget_rows), 2)
    if abs(actual_revenue - budget_revenue_total) > 0.05:
        errors.append(f"{path.name} revenue detail does not reconcile to Budget_vs_Actual")
    if abs(actual_expense - budget_expense_total) > 0.05:
        errors.append(f"{path.name} expense detail does not reconcile to Budget_vs_Actual")
    if payroll_total <= 0 or payroll_total > actual_expense:
        errors.append(f"{path.name} payroll total is invalid relative to expenses")
    if cash_rows:
        beginning = float(cash_rows[0].get("Beginning_Cash") or 0.0)
        ending = float(cash_rows[0].get("Actual_Ending_Cash") or 0.0)
        if round(beginning + cash_flow, 2) != round(ending, 2):
            errors.append(f"{path.name} cash flow does not reconcile to ending cash")

    return {
        "errors": errors,
        "period": period,
        "totals": {
            "actual_revenue": actual_revenue,
            "budget_revenue": budget_revenue,
            "actual_expense": actual_expense,
            "budget_expense": budget_expense,
            "payroll_total": payroll_total,
            "payroll_ratio": round(payroll_total / actual_revenue if actual_revenue else 0.0, 4),
            "net_cash_flow": cash_flow,
        },
    }


def _read_sheet_rows(wb: Any, sheet_name: str) -> list[dict[str, Any]]:
    """Read a generated workbook sheet into dictionaries.

    Inputs:
        wb: OpenPyXL workbook.
        sheet_name: Sheet to read.
    Outputs:
        Row dictionaries keyed by header row 5.
    Assumptions:
        Empty first cell after the header indicates no more data rows.
    """

    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [value for value in next(ws.iter_rows(min_row=5, max_row=5, values_only=True)) if value is not None]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=6, values_only=True):
        if values[0] is None:
            break
        rows.append({header: values[idx] for idx, header in enumerate(headers)})
    return rows


def _period_from_path(path: Path) -> str:
    """Extract the period slug from a generated workbook filename.

    Inputs:
        path: Workbook path.
    Outputs:
        Period slug such as ``2026_06``.
    Assumptions:
        Generated report names end with ``YYYY_MM.xlsx``.
    """

    stem_parts = path.stem.split("_")
    return f"{stem_parts[-2]}_{stem_parts[-1]}"


def _validate_manifest_patterns(manifest: dict[str, Any], workbook_periods: list[str], errors: list[str]) -> None:
    """Validate scenario-level manifest trends and anomaly periods.

    Inputs:
        manifest: Scenario manifest dictionary.
        workbook_periods: Periods observed in workbooks.
        errors: Mutable error list.
    Outputs:
        None.
    Assumptions:
        The recovery scenario encodes a payroll peak in June and collection
        recovery after August.
    """

    if workbook_periods != sorted(workbook_periods):
        errors.append("Workbook periods are not chronological")
    payroll_trend = manifest.get("monthly_payroll_ratio_trend", {})
    collection_trend = manifest.get("collection_rate_trend", {})
    if payroll_trend:
        peak_period = max(payroll_trend, key=lambda key: payroll_trend[key])
        if not peak_period.endswith("_06"):
            errors.append(f"Expected payroll ratio peak in June, found {peak_period}")
    if collection_trend:
        if float(collection_trend.get("2026_08", 0.0)) <= float(collection_trend.get("2026_07", 0.0)):
            errors.append("Expected collection campaign improvement in August")
        if float(collection_trend.get("2026_12", 0.0)) <= float(collection_trend.get("2026_06", 0.0)):
            errors.append("Expected year-end collection recovery versus June")
    if manifest.get("health_sciences_overspending_periods") != ["2026_04", "2026_05", "2026_06", "2026_07"]:
        errors.append("Health Sciences overspending periods do not match recovery scenario")
    if manifest.get("recurring_vendor_anomaly_periods") != ["2026_07", "2026_08", "2026_09"]:
        errors.append("Recurring vendor anomaly periods do not match recovery scenario")
